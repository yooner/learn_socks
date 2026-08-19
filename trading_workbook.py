"""Excel trading-management workbook generator."""

from __future__ import annotations

import math
from datetime import date, timedelta
from pathlib import Path
from typing import Any, Iterable, Mapping

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableColumn, TableStyleInfo

TRADE_HEADERS = [
    "交易编号",
    "股票代码",
    "买入时账户金额",
    "本次允许亏损比例",
    "买入价",
    "买入建议股数",
    "开仓风险告警",
    "实际买入股数",
    "买入日期",
    "期望卖出价",
    "实际卖出价",
    "卖出股数",
    "卖出日期",
    "止损价",
    "买入费用",
    "卖出费用",
    "买入价的由来",
    "止损价的由来",
    "期望卖出价的由来",
    "实际卖出价的由来",
    "期望盈利比例",
    "期望止损比例",
    "盈亏比",
    "实际盈亏金额",
    "单笔仓位收益率",
    "账户收益率与平均盈利率差值",
    "持有天数",
    "复利容许平均亏损上限",
    "复利风险判断",
    "交易打分评价",
    "实际账户收益率",
]

REASON_HEADERS = [
    "交易编号",
    "股票代码",
    "所属板块",
    "阶段",
    "记录日期",
    "技术指标1",
    "技术指标2",
    "技术指标3",
    "综述",
]

TRACKING_HEADERS = [
    "记录类型",
    "交易编号",
    "股票代码",
    "止损阶段序号",
    "激活价",
    "激活后止损价",
    "止损依据",
    "记录日期",
    "收盘价",
    "成交量",
    "量价分析",
    "当前阶段",
    "阶段判断理由",
    "初始止损",
    "历史最高收盘价",
    "已激活止损阶段",
    "当前计划止损",
    "人工上调止损",
    "当前有效止损",
    "下一阶段激活价",
    "下一阶段止损价",
    "距离下一阶段涨幅",
    "止损判断",
    "卖出动作",
    "实际卖出价",
    "实际卖出股数",
    "卖出费用",
    "执行检查",
    "规则检查",
]

TRADE_EXPECTATION_TYPES = ("短期博反弹", "突破冲新高", "趋势波段")
HOLDING_PERIODS = (
    "1～3个交易日",
    "4～10个交易日",
    "11～20个交易日",
    "21～60个交易日",
    "60个交易日以上",
)
EXPECTATION_PERIOD_MATCHES = {
    "短期博反弹": {"1～3个交易日", "4～10个交易日"},
    "突破冲新高": {"4～10个交易日", "11～20个交易日"},
    "趋势波段": {"11～20个交易日", "21～60个交易日", "60个交易日以上"},
}


def calculate_tranche_position(
    tranches: Iterable[tuple[float | None, int | None]],
    effective_stop: float | None,
    is_closed: bool = False,
) -> dict[str, float | int | None]:
    """Calculate aggregate cost, shares, weighted price, and open risk."""
    valid = [
        (float(price), int(shares))
        for price, shares in tranches
        if price is not None and shares is not None and shares > 0
    ]
    total_shares = sum(shares for _, shares in valid)
    buy_amount = sum(price * shares for price, shares in valid)
    weighted_buy_price = (
        buy_amount / total_shares if total_shares else None
    )
    stop = float(effective_stop) if effective_stop is not None else None
    current_risk = (
        0.0
        if is_closed or stop is None
        else sum(
            max(price - stop, 0.0) * shares
            for price, shares in valid
        )
    )
    return {
        "total_shares": total_shares,
        "buy_amount": buy_amount,
        "weighted_buy_price": weighted_buy_price,
        "current_risk": current_risk,
    }


def check_tranche_rules(
    trade_id: str | None,
    tranches: Iterable[tuple[float | None, int | None]],
    is_locked: bool,
    position_risk: float | None,
    single_trade_limit: float | None,
    total_open_risk: float | None,
    account_risk_limit: float | None,
) -> str:
    """Return the first violated three-batch position rule."""
    if not trade_id:
        return ""
    batches = list(tranches)
    if len(batches) != 3:
        raise ValueError("exactly three tranche slots are required")

    first_price, first_shares = batches[0]
    if first_price is None or first_shares is None:
        return "违规：第一批价格或股数缺失"

    for index, (price, shares) in enumerate(batches[1:], start=2):
        if (price is None) != (shares is None):
            return f"违规：第{'二' if index == 2 else '三'}批价格与股数须成对填写"

    second_complete = all(value is not None for value in batches[1])
    third_complete = all(value is not None for value in batches[2])
    if third_complete and not second_complete:
        return "违规：必须先完成第二批"

    completed_shares = [
        shares
        for price, shares in batches
        if price is not None and shares is not None
    ]
    if any(
        not isinstance(shares, (int, float))
        or shares <= 0
        or shares != int(shares)
        or int(shares) % 100 != 0
        for shares in completed_shares
    ):
        return "违规：股数须为100股整数倍"

    has_addition = second_complete or third_complete
    if has_addition and is_locked:
        return "违规：锁仓期间禁止加仓"
    if (
        position_risk is not None
        and single_trade_limit is not None
        and position_risk > single_trade_limit
    ):
        return "违规：超过单笔风险上限"
    if (
        total_open_risk is not None
        and account_risk_limit is not None
        and total_open_risk > account_risk_limit
    ):
        return "违规：超过账户风险上限"
    return "通过"

TECHNICAL_INDICATORS = [
    ("蜡烛图", "价格形态", "观察实体、上下影线及组合形态"),
    ("趋势线", "趋势", "连接关键高点或低点判断趋势方向"),
    ("MACD", "动量", "观察快慢线、信号线和柱状图"),
    ("移动平均线", "趋势", "观察价格与不同周期均线的位置关系"),
    ("成交量", "量价", "确认突破、回调和趋势的参与强度"),
    ("RSI", "动量", "衡量相对强弱及超买超卖状态"),
    ("KDJ", "动量", "观察随机指标交叉和极值区域"),
    ("布林带", "波动", "观察价格相对中轨和上下轨的位置"),
    ("支撑位", "关键价位", "记录预期获得买盘支撑的价格区域"),
    ("压力位", "关键价位", "记录预期遇到卖压的价格区域"),
    ("缺口", "价格形态", "观察跳空缺口是否回补或延续"),
    ("形态突破", "价格形态", "观察平台、箱体或整理形态的突破"),
    ("均线金叉", "趋势", "短周期均线上穿长周期均线"),
    ("均线死叉", "趋势", "短周期均线下穿长周期均线"),
    ("量价背离", "量价", "价格和成交量变化方向不一致"),
    ("趋势背离", "动量", "价格走势与动量指标走势不一致"),
    ("ATR", "波动", "衡量真实波动幅度并辅助设置止损"),
    ("换手率", "量价", "衡量筹码交换活跃程度"),
    ("相对强弱", "比较", "比较个股与板块或指数的强弱"),
    ("板块共振", "市场环境", "确认个股信号与所属板块方向一致"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
INPUT_FILL = PatternFill("solid", fgColor="DDEBF7")
FORMULA_FILL = PatternFill("solid", fgColor="E7E6E6")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
YELLOW_FILL = PatternFill("solid", fgColor="FFEB9C")
RED_FILL = PatternFill("solid", fgColor="FFC7CE")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
CURRENCY_FORMAT = '¥#,##0.00;[Red]-¥#,##0.00'
PERCENT_FORMAT = "0.00%"


def calculate_position_size(
    account_snapshot: float,
    risk_rate: float,
    buy_price: float,
    stop_price: float,
) -> int | None:
    """Return the risk-sized position rounded down to a 100-share lot."""
    if (
        account_snapshot <= 0
        or risk_rate <= 0
        or buy_price <= 0
        or stop_price <= 0
        or buy_price <= stop_price
    ):
        return None
    raw_shares = account_snapshot * risk_rate / (buy_price - stop_price)
    return math.floor(raw_shares / 100) * 100


def calculate_realized_pnl(
    buy_price: float,
    buy_shares: int,
    sell_price: float,
    sell_shares: int,
    buy_fee: float = 0,
    sell_fee: float = 0,
) -> float:
    """Calculate realized profit after explicit buy and sell fees."""
    return (
        sell_price * sell_shares
        - buy_price * buy_shares
        - buy_fee
        - sell_fee
    )


def calculate_return_rate(
    pnl: float,
    buy_price: float,
    buy_shares: int,
    buy_fee: float = 0,
) -> float | None:
    """Calculate return on the cash committed to the purchase."""
    invested = buy_price * buy_shares + buy_fee
    if invested <= 0:
        return None
    return pnl / invested


def calculate_account_return(
    pnl: float,
    account_snapshot: float,
) -> float | None:
    """Calculate one trade's realized contribution to total account equity."""
    if account_snapshot <= 0:
        return None
    return pnl / account_snapshot


def calculate_compound_return(
    account_returns: Iterable[float | None],
) -> float | None:
    """Compound actual account returns in transaction order."""
    factor = 1.0
    count = 0
    for account_return in account_returns:
        if account_return is None:
            continue
        factor *= 1 + account_return
        count += 1
    return factor - 1 if count else None


def calculate_open_theoretical_loss(
    trades: Iterable[Mapping[str, Any]],
) -> float:
    """Return stop-loss exposure for positions without a sell date."""
    total = 0.0
    for trade in trades:
        if trade.get("sell_date") is not None:
            continue
        buy_price = trade.get("buy_price")
        stop_price = trade.get("stop_price")
        actual_shares = trade.get("actual_buy_shares")
        if (
            buy_price is None
            or stop_price is None
            or actual_shares is None
            or actual_shares <= 0
        ):
            continue
        loss_per_share = float(buy_price) - float(stop_price)
        if loss_per_share > 0:
            total += loss_per_share * int(actual_shares)
    return total


def calculate_opening_risk_status(
    open_risk: float,
    candidate_risk: float,
    monthly_limit: float | None,
) -> str | None:
    """Return whether a candidate position fits inside the monthly risk limit."""
    if monthly_limit is None or monthly_limit <= 0:
        return None
    if open_risk + candidate_risk >= monthly_limit:
        return "禁止开仓"
    return "允许开仓"


def _average(values: Iterable[float]) -> float | None:
    items = list(values)
    if not items:
        return None
    return sum(items) / len(items)


def summarize_trades(
    trades: Iterable[Mapping[str, Any]],
) -> dict[str, float | int | None]:
    """Aggregate closed-trade statistics using the agreed metric definitions."""
    items = list(trades)
    closed = [
        trade
        for trade in items
        if trade.get("sell_date") is not None and trade.get("pnl") is not None
    ]
    wins = [trade for trade in closed if trade["pnl"] > 0]
    losses = [trade for trade in closed if trade["pnl"] < 0]
    flats = [trade for trade in closed if trade["pnl"] == 0]
    decisive_count = len(wins) + len(losses)
    win_rate = len(wins) / decisive_count if decisive_count else None
    loss_rate = len(losses) / decisive_count if decisive_count else None
    average_win = _average(
        trade["account_return"]
        for trade in wins
        if trade.get("account_return") is not None
    )
    average_loss = _average(
        abs(trade["account_return"])
        for trade in losses
        if trade.get("account_return") is not None
    )
    expectancy = None
    if (
        win_rate is not None
        and loss_rate is not None
        and average_win is not None
        and average_loss is not None
    ):
        expectancy = win_rate * average_win - loss_rate * average_loss

    compound_return = calculate_compound_return(
        trade.get("account_return") for trade in closed
    )

    return {
        "completed_count": len(closed),
        "win_count": len(wins),
        "loss_count": len(losses),
        "flat_count": len(flats),
        "win_rate": win_rate,
        "loss_rate": loss_rate,
        "average_win": average_win,
        "average_loss": average_loss,
        "win_hold_days": sum(trade.get("hold_days") or 0 for trade in wins),
        "loss_hold_days": sum(trade.get("hold_days") or 0 for trade in losses),
        "average_trade_amount": _average(
            trade["trade_amount"]
            for trade in items
            if trade.get("trade_amount") is not None
            and trade["trade_amount"] > 0
        ),
        "expectancy": expectancy,
        "compound_return": compound_return,
    }


def calculate_compound_loss_ceiling(
    average_profit: float | None,
    win_rate: float | None,
) -> float | None:
    """Return the average-loss ceiling implied by geometric break-even."""
    if (
        average_profit is None
        or average_profit <= -1
        or win_rate is None
        or not 0 < win_rate < 1
    ):
        return None
    return 1 - (1 + average_profit) ** (-win_rate / (1 - win_rate))


def calculate_monthly_loss(
    trades: Iterable[Mapping[str, Any]],
    as_of_date: date,
) -> float:
    """Sum current-month realized losses while preserving their negative sign."""
    total = 0.0
    for trade in trades:
        sell_date = trade.get("sell_date")
        pnl = trade.get("pnl")
        if (
            isinstance(sell_date, date)
            and sell_date.year == as_of_date.year
            and sell_date.month == as_of_date.month
            and pnl is not None
            and pnl < 0
        ):
            total += pnl
    return total


def calculate_consecutive_loss_lock(
    pnls: Iterable[float | None],
    initial_balance: float,
    loss_limit_rate: float = 0.06,
    manual_unlock_through: int | None = None,
    unlock_reason: str | None = None,
) -> dict[str, float | int | str]:
    """Evaluate the loss cycle after the latest profitable record.

    Record sequences are one-based. A valid manual unlock closes the current
    cycle through the selected record; the next record starts a fresh cycle.
    """
    items = list(pnls)
    latest_closed_sequence = max(
        (index for index, pnl in enumerate(items, start=1) if pnl is not None),
        default=0,
    )
    latest_win_sequence = max(
        (
            index
            for index, pnl in enumerate(items, start=1)
            if pnl is not None and pnl > 0
        ),
        default=0,
    )
    has_unlock_sequence = manual_unlock_through is not None
    has_unlock_reason = bool(unlock_reason and unlock_reason.strip())
    unlock_is_valid = (
        has_unlock_sequence
        and has_unlock_reason
        and isinstance(manual_unlock_through, int)
        and 1 <= manual_unlock_through <= latest_closed_sequence
    )
    unlock_is_incomplete = (
        has_unlock_sequence != has_unlock_reason
        or (has_unlock_sequence and has_unlock_reason and not unlock_is_valid)
    )
    cycle_start_sequence = max(
        latest_win_sequence,
        manual_unlock_through if unlock_is_valid else 0,
    )
    cycle_start_balance = initial_balance + sum(
        pnl or 0 for pnl in items[:cycle_start_sequence]
    )
    cycle_loss = -sum(
        pnl
        for pnl in items[cycle_start_sequence:]
        if pnl is not None and pnl < 0
    )
    loss_ratio = (
        cycle_loss / cycle_start_balance
        if cycle_start_balance > 0
        else 0.0
    )
    risk_usage = (
        loss_ratio / loss_limit_rate if loss_limit_rate > 0 else 0.0
    )
    if unlock_is_incomplete:
        status = "解锁信息不完整"
    elif risk_usage >= 1:
        status = "已锁仓"
    elif risk_usage >= 0.8:
        status = "接近锁仓"
    else:
        status = "正常"
    return {
        "latest_win_sequence": latest_win_sequence,
        "latest_closed_sequence": latest_closed_sequence,
        "cycle_start_sequence": cycle_start_sequence,
        "cycle_start_balance": cycle_start_balance,
        "cycle_loss": cycle_loss,
        "loss_ratio": loss_ratio,
        "risk_usage": risk_usage,
        "status": status,
    }


def calculate_dynamic_stop_history(
    initial_stop: float,
    plan_levels: Iterable[Mapping[str, float | int]],
    closing_prices: Iterable[float],
    manual_stops: Iterable[float | None] | None = None,
) -> list[dict[str, float | int | str | bool | None]]:
    """Calculate long-position stop states from daily closing prices."""
    levels = sorted(
        (
            {
                "stage": int(level["stage"]),
                "activation_price": float(level["activation_price"]),
                "stop_price": float(level["stop_price"]),
            }
            for level in plan_levels
        ),
        key=lambda level: (level["activation_price"], level["stage"]),
    )
    closes = [float(price) for price in closing_prices]
    manual_values = (
        list(manual_stops)
        if manual_stops is not None
        else [None] * len(closes)
    )
    if len(manual_values) != len(closes):
        raise ValueError("manual_stops must match closing_prices length")

    historical_high = float("-inf")
    effective_stop = float(initial_stop)
    result: list[dict[str, float | int | str | bool | None]] = []
    for close, manual_stop in zip(closes, manual_values):
        historical_high = max(historical_high, close)
        activated = [
            level
            for level in levels
            if historical_high >= level["activation_price"]
        ]
        activated_stage = max(
            (level["stage"] for level in activated),
            default=0,
        )
        plan_stop = max(
            [float(initial_stop)]
            + [level["stop_price"] for level in activated]
        )
        effective_stop = max(effective_stop, plan_stop)
        manual_stop_rejected = False
        if manual_stop is not None:
            manual_value = float(manual_stop)
            if manual_value < effective_stop:
                manual_stop_rejected = True
            else:
                effective_stop = manual_value
        pending = [
            level
            for level in levels
            if level["activation_price"] > historical_high
        ]
        next_level = min(
            pending,
            key=lambda level: (level["activation_price"], level["stage"]),
            default=None,
        )
        result.append(
            {
                "closing_price": close,
                "historical_high": historical_high,
                "activated_stage": activated_stage,
                "plan_stop": plan_stop,
                "effective_stop": effective_stop,
                "next_activation_price": (
                    next_level["activation_price"] if next_level else None
                ),
                "next_stop_price": (
                    next_level["stop_price"] if next_level else None
                ),
                "manual_stop_rejected": manual_stop_rejected,
                "stop_status": (
                    "触发止损：应全部卖出"
                    if close <= effective_stop
                    else "继续观察"
                ),
            }
        )
    return result


def check_trade_expectation_period(
    trade_type: str | None,
    holding_period: str | None,
) -> str | None:
    """Return whether a selected holding range matches the trade thesis."""
    if not trade_type or not holding_period:
        return None
    if holding_period in EXPECTATION_PERIOD_MATCHES.get(trade_type, set()):
        return "匹配"
    return "周期与交易预期不匹配，请重新确认"


def calculate_required_trades(
    target_profit: float,
    average_trade_amount: float,
    expected_return_rate: float,
) -> int | None:
    """Return whole trades required to reach a target, rounded upward."""
    if (
        target_profit <= 0
        or average_trade_amount <= 0
        or expected_return_rate <= 0
    ):
        return None
    return math.ceil(
        target_profit / (average_trade_amount * expected_return_rate)
    )


def _apply_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = "A2"


def _add_table(ws, name: str, ref: str) -> None:
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _add_validation(
    ws,
    validation: DataValidation,
    cell_range: str,
) -> None:
    ws.add_data_validation(validation)
    for range_part in cell_range.split():
        validation.add(range_part)


def _style_trade_sheet(ws, end_row: int) -> None:
    input_columns = {
        1,
        2,
        3,
        4,
        5,
        8,
        9,
        10,
        11,
        13,
        14,
        15,
        16,
        17,
        18,
        19,
        20,
        30,
    }
    for row in range(2, end_row + 1):
        for column in range(1, len(TRADE_HEADERS) + 1):
            cell = ws.cell(row, column)
            cell.fill = (
                INPUT_FILL if column in input_columns else FORMULA_FILL
            )
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in (3, 5, 10, 11, 14, 15, 16, 24):
            ws.cell(row, column).number_format = CURRENCY_FORMAT
        for column in (4, 21, 22, 25, 26, 28, 31):
            ws.cell(row, column).number_format = PERCENT_FORMAT
        for column in (9, 13):
            ws.cell(row, column).number_format = "yyyy-mm-dd"
        ws.cell(row, 23).number_format = "0.00"
        ws.cell(row, 27).number_format = "0"

    width_by_header = {
        "交易编号": 14,
        "股票代码": 14,
        "买入时账户金额": 17,
        "本次允许亏损比例": 18,
        "买入建议股数": 16,
        "开仓风险告警": 16,
        "实际买入股数": 16,
        "买入日期": 13,
        "卖出日期": 13,
        "买入价的由来": 24,
        "止损价的由来": 24,
        "期望卖出价的由来": 24,
        "实际卖出价的由来": 24,
        "复利容许平均亏损上限": 22,
        "复利风险判断": 18,
        "交易打分评价": 18,
        "实际账户收益率": 18,
    }
    for column, header in enumerate(TRADE_HEADERS, start=1):
        ws.column_dimensions[get_column_letter(column)].width = (
            width_by_header.get(header, 14)
        )


def _add_trade_formulas(ws, end_row: int) -> None:
    for row in range(2, end_row + 1):
        ws.cell(
            row,
            6,
            (
                f'=IF(OR(C{row}="",D{row}="",E{row}="",N{row}="",'
                f"C{row}<=0,D{row}<=0,E{row}<=N{row}),"
                f'"",ROUNDDOWN((C{row}*D{row})/(E{row}-N{row})/100,0)*100)'
            ),
        )
        ws.cell(
            row,
            7,
            (
                f'=IF(OR(A{row}="",F{row}="",E{row}="",N{row}="",'
                f"E{row}<=N{row},'账户数据'!$B$7=\"\"),\"\","
                f"IF('账户数据'!$B$9+IF(H{row}=\"\","
                f"F{row}*(E{row}-N{row}),0)>='账户数据'!$B$7,"
                '"禁止开仓","允许开仓"))'
            ),
        )
        ws.cell(
            row,
            12,
            f'=IF(K{row}="","",H{row})',
        )
        ws.cell(
            row,
            21,
            f'=IF(OR(E{row}="",J{row}="",E{row}<=0),"",(J{row}-E{row})/E{row})',
        )
        ws.cell(
            row,
            22,
            f'=IF(OR(E{row}="",N{row}="",E{row}<=0),"",(E{row}-N{row})/E{row})',
        )
        ws.cell(
            row,
            23,
            f'=IF(OR(U{row}="",V{row}="",V{row}<=0),"",U{row}/V{row})',
        )
        ws.cell(
            row,
            24,
            (
                f'=IF(OR(K{row}="",L{row}="",E{row}="",H{row}=""),"",'
                f'K{row}*L{row}-E{row}*H{row}-IF(O{row}="",0,O{row})'
                f'-IF(P{row}="",0,P{row}))'
            ),
        )
        ws.cell(
            row,
            25,
            (
                f'=IF(OR(X{row}="",E{row}="",H{row}=""),"",'
                f'IF(E{row}*H{row}+IF(O{row}="",0,O{row})<=0,"",'
                f'X{row}/(E{row}*H{row}+IF(O{row}="",0,O{row}))))'
            ),
        )
        ws.cell(
            row,
            26,
            (
                f'=IF(OR(AE{row}="",\'多次统计数据\'!$B$8=""),"",'
                f"AE{row}-'多次统计数据'!$B$8)"
            ),
        )
        ws.cell(
            row,
            27,
            f'=IF(OR(I{row}="",M{row}="",M{row}<I{row}),"",M{row}-I{row})',
        )
        ws.cell(
            row,
            28,
            (
                f'=IF(OR(A{row}="",\'多次统计数据\'!$B$8="",'
                "'多次统计数据'!$B$6<=0,"
                "'多次统计数据'!$B$6>=1),\"\","
                "1-POWER(1+'多次统计数据'!$B$8,"
                "-'多次统计数据'!$B$6/"
                "(1-'多次统计数据'!$B$6)))"
            ),
        )
        ws.cell(
            row,
            29,
            (
                f'=IF(OR(AB{row}="",\'多次统计数据\'!$B$9=""),"",'
                f'IF(\'多次统计数据\'!$B$9<AB{row},"低于上限",'
                f'"达到或超过上限"))'
            ),
        )
        ws.cell(
            row,
            31,
            f'=IF(OR(X{row}="",C{row}="",C{row}<=0),"",X{row}/C{row})',
        )


def _add_trade_validations(ws, end_row: int) -> None:
    positive_decimal = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
    )
    _add_validation(
        ws,
        positive_decimal,
        f"C2:C{end_row} E2:E{end_row} J2:K{end_row} N2:N{end_row}",
    )
    risk_rate = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=True,
    )
    _add_validation(ws, risk_rate, f"D2:D{end_row}")
    nonnegative_fee = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    _add_validation(ws, nonnegative_fee, f"O2:P{end_row}")
    actual_shares = DataValidation(
        type="custom",
        formula1='=OR(H2="",AND(ISNUMBER(H2),H2>0,MOD(H2,100)=0))',
        allow_blank=True,
    )
    actual_shares.error = "实际买入股数必须是大于0的100股整数倍"
    actual_shares.errorTitle = "实际买入股数无效"
    actual_shares.showErrorMessage = True
    _add_validation(ws, actual_shares, f"H2:H{end_row}")
    buy_date = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2000,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    _add_validation(ws, buy_date, f"I2:I{end_row}")
    sell_date = DataValidation(
        type="custom",
        formula1='=OR(M2="",AND(ISNUMBER(M2),M2>=I2))',
        allow_blank=True,
    )
    _add_validation(ws, sell_date, f"M2:M{end_row}")
    sell_shares = DataValidation(
        type="custom",
        formula1='=OR(L2="",L2=H2)',
        allow_blank=True,
    )
    _add_validation(ws, sell_shares, f"L2:L{end_row}")
    score = DataValidation(
        type="list",
        formula1='"1-差,2-较差,3-一般,4-良好,5-优秀"',
        allow_blank=True,
    )
    _add_validation(ws, score, f"AD2:AD{end_row}")


def _add_trade_conditional_formatting(ws, end_row: int) -> None:
    ws.conditional_formatting.add(
        f"X2:X{end_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        f"X2:X{end_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        f"W2:W{end_row}",
        CellIsRule(operator="lessThan", formula=["2"], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        f"AC2:AC{end_row}",
        FormulaRule(
            formula=['$AC2="达到或超过上限"'],
            fill=RED_FILL,
        ),
    )
    ws.conditional_formatting.add(
        f"AC2:AC{end_row}",
        FormulaRule(formula=['$AC2="低于上限"'], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        f"Y2:Y{end_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        f"Y2:Y{end_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        f"AE2:AE{end_row}",
        CellIsRule(operator="greaterThan", formula=["0"], fill=GREEN_FILL),
    )
    ws.conditional_formatting.add(
        f"AE2:AE{end_row}",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        f"G2:G{end_row}",
        FormulaRule(formula=['$G2="禁止开仓"'], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        f"G2:G{end_row}",
        FormulaRule(formula=['$G2="允许开仓"'], fill=GREEN_FILL),
    )


def _build_trade_sheet(wb: Workbook, end_row: int = 201):
    ws = wb.create_sheet("单次交易")
    _apply_header(ws, TRADE_HEADERS)
    _add_trade_formulas(ws, end_row)
    _style_trade_sheet(ws, end_row)
    _add_trade_validations(ws, end_row)
    _add_trade_conditional_formatting(ws, end_row)
    _add_table(ws, "TradeRecords", f"A1:AE{end_row}")
    ws["C1"].comment = Comment(
        "开仓时从“账户数据”的当前总金额复制，并粘贴为数值；不要保留公式。",
        "Codex",
    )
    ws["D1"].comment = Comment(
        "开仓时从“账户数据”的单次交易允许亏损比例复制，并粘贴为数值。",
        "Codex",
    )
    ws["F1"].comment = Comment(
        "按账户快照、本次风险比例和止损距离计算；仅作仓位建议。",
        "Codex",
    )
    ws["G1"].comment = Comment(
        "比较全部未平仓理论亏损（含本行拟开仓风险）与当月允许最大亏损金额。",
        "Codex",
    )
    ws["H1"].comment = Comment(
        "手动填写实际成交股数；后续盈亏、收益率和统计均引用此列。",
        "Codex",
    )
    ws["L1"].comment = Comment(
        "默认等于实际买入股数；本系统按一次卖出处理。",
        "Codex",
    )
    ws["AE1"].comment = Comment(
        "实际盈亏金额÷买入时账户金额快照；用于账户口径统计与逐笔复利。",
        "Codex",
    )
    return ws


def _build_reason_sheet(wb: Workbook, end_row: int = 501):
    ws = wb.create_sheet("买入理由")
    _apply_header(ws, REASON_HEADERS)
    for row in range(2, end_row + 1):
        for column in range(1, len(REASON_HEADERS) + 1):
            cell = ws.cell(row, column)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, 5).number_format = "yyyy-mm-dd"
    widths = [14, 14, 16, 14, 13, 24, 24, 24, 48]
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    stage = DataValidation(
        type="list",
        formula1='"买入,持续追踪,卖出"',
        allow_blank=True,
    )
    _add_validation(ws, stage, f"D2:D{end_row}")
    log_date = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2000,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    _add_validation(ws, log_date, f"E2:E{end_row}")
    indicator = DataValidation(
        type="list",
        formula1="=技术指标列表",
        allow_blank=True,
    )
    indicator.error = "请从“技术指标”工作表维护的列表中选择"
    indicator.errorTitle = "技术指标无效"
    indicator.showErrorMessage = True
    _add_validation(ws, indicator, f"F2:H{end_row}")
    _add_table(ws, "TradeReasonLog", f"A1:I{end_row}")
    return ws


def _tracking_row_formulas(row: int, end_row: int) -> dict[int, str]:
    daily = "每日跟踪"
    plan = "止损计划"
    formulas = {
        3: (
            f'=IF(B{row}="","",IFERROR(LOOKUP(2,1/'
            f"('单次交易'!$A$2:$A$201=B{row}),"
            "'单次交易'!$B$2:$B$201),\"\"))"
        ),
        14: (
            f'=IF(B{row}="","",IFERROR(LOOKUP(2,1/'
            f"('单次交易'!$A$2:$A$201=B{row}),"
            "'单次交易'!$N$2:$N$201),\"\"))"
        ),
        15: (
            f'=IF(A{row}<>"{daily}","",IF(OR(B{row}="",H{row}="",I{row}=""),"",'
            f'MAXIFS($I$2:$I${end_row},$A$2:$A${end_row},"{daily}",'
            f'$B$2:$B${end_row},B{row},$H$2:$H${end_row},"<="&H{row})))'
        ),
        16: (
            f'=IF(A{row}<>"{daily}","",IF(O{row}="","",IFERROR('
            f'MAXIFS($D$2:$D${end_row},$A$2:$A${end_row},"{plan}",'
            f'$B$2:$B${end_row},B{row},$E$2:$E${end_row},"<="&O{row}),0)))'
        ),
        17: (
            f'=IF(A{row}<>"{daily}","",IF(N{row}="","",MAX(N{row},IFERROR('
            f'MAXIFS($F$2:$F${end_row},$A$2:$A${end_row},"{plan}",'
            f'$B$2:$B${end_row},B{row},$E$2:$E${end_row},"<="&O{row}),0))))'
        ),
        19: (
            f'=IF(A{row}<>"{daily}","",IF(Q{row}="","",MAX(Q{row},IFERROR('
            f'MAXIFS($R$2:$R${end_row},$A$2:$A${end_row},"{daily}",'
            f'$B$2:$B${end_row},B{row},$H$2:$H${end_row},"<="&H{row}),0))))'
        ),
        20: (
            f'=IF(A{row}<>"{daily}","",IF(O{row}="","",IF('
            f'COUNTIFS($A$2:$A${end_row},"{plan}",$B$2:$B${end_row},B{row},'
            f'$E$2:$E${end_row},">"&O{row})=0,"",'
            f'MINIFS($E$2:$E${end_row},$A$2:$A${end_row},"{plan}",'
            f'$B$2:$B${end_row},B{row},$E$2:$E${end_row},">"&O{row}))))'
        ),
        21: (
            f'=IF(OR(A{row}<>"{daily}",T{row}=""),"",IFERROR('
            f'MAXIFS($F$2:$F${end_row},$A$2:$A${end_row},"{plan}",'
            f'$B$2:$B${end_row},B{row},$E$2:$E${end_row},T{row}),""))'
        ),
        22: (
            f'=IF(OR(A{row}<>"{daily}",T{row}="",I{row}="",I{row}<=0),"",'
            f'T{row}/I{row}-1)'
        ),
        23: (
            f'=IF(A{row}<>"{daily}","",IF(OR(I{row}="",S{row}=""),"",'
            f'IF(I{row}<=S{row},"触发止损：应全部卖出","继续观察")))'
        ),
        26: (
            f'=IF(X{row}<>"执行卖出","",IFERROR(LOOKUP(2,1/'
            f"('单次交易'!$A$2:$A$201=B{row}),"
            "'单次交易'!$H$2:$H$201),\"\"))"
        ),
        28: (
            f'=IF(A{row}<>"{daily}","",IF(OR(B{row}="",H{row}="",I{row}="",'
            f'J{row}="",K{row}="",L{row}="",M{row}=""),"待补充每日记录",'
            f'IF(W{row}="触发止损：应全部卖出",IF(X{row}<>"执行卖出",'
            f'"违规：触发止损但未执行",IF(Y{row}="","待补充实际卖出价",'
            f'"卖出记录完整")),IF(X{row}="","待选择卖出动作",'
            f'IF(X{row}="执行卖出",IF(Y{row}="","待补充实际卖出价",'
            f'"卖出记录完整"),"正常")))))'
        ),
    }
    plan_check = (
        f'IF(OR(B{row}="",D{row}="",E{row}="",F{row}="",G{row}=""),"计划不完整",'
        f'IF(OR(D{row}<1,D{row}<>INT(D{row})),"阶段序号无效",'
        f'IF(COUNTIFS($A$2:$A${end_row},"{plan}",$B$2:$B${end_row},B{row},'
        f'$D$2:$D${end_row},D{row})>1,"阶段序号重复",'
        f'IF(E{row}<=IFERROR(LOOKUP(2,1/(\'单次交易\'!$A$2:$A$201=B{row}),'
        "'单次交易'!$E$2:$E$201),0),\"激活价须高于买入价\","
        f'IF(F{row}<N{row},"止损价不得低于初始止损",'
        f'IF(F{row}>E{row},"止损价不得高于激活价",'
        f'IF(AND(COUNTIFS($A$2:$A${end_row},"{plan}",$B$2:$B${end_row},B{row},'
        f'$D$2:$D${end_row},"<"&D{row})>0,OR('
        f'E{row}<=MAXIFS($E$2:$E${end_row},$A$2:$A${end_row},"{plan}",'
        f'$B$2:$B${end_row},B{row},$D$2:$D${end_row},"<"&D{row}),'
        f'F{row}<MAXIFS($F$2:$F${end_row},$A$2:$A${end_row},"{plan}",'
        f'$B$2:$B${end_row},B{row},$D$2:$D${end_row},"<"&D{row}))),'
        '"阶段价格必须逐级提高","计划有效")))))))'
    )
    daily_check = (
        f'IF(OR(B{row}="",H{row}="",I{row}="",J{row}=""),"每日数据不完整",'
        f'IF(COUNTIFS($A$2:$A${end_row},"{daily}",$B$2:$B${end_row},B{row},'
        f'$H$2:$H${end_row},H{row})>1,"同一交易日期重复",'
        f'IF(AND(R{row}<>"",R{row}<S{row}),"人工止损无效：不能下调","跟踪有效")))'
    )
    formulas[29] = (
        f'=IF(A{row}="","",IF(A{row}="{plan}",{plan_check},'
        f'IF(A{row}="{daily}",{daily_check},"记录类型无效")))'
    )
    return formulas


def _build_tracking_sheet(
    wb: Workbook,
    end_row: int = 501,
):
    ws = wb.create_sheet("持仓跟踪", 1)
    _apply_header(ws, TRACKING_HEADERS)
    ws.freeze_panes = "D2"
    manual_columns = {1, 2, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 18, 24, 25, 27}
    for row in range(2, end_row + 1):
        formulas = _tracking_row_formulas(row, end_row)
        for column in range(1, len(TRACKING_HEADERS) + 1):
            cell = ws.cell(row, column)
            if column in formulas:
                cell.value = formulas[column]
            cell.fill = INPUT_FILL if column in manual_columns else FORMULA_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.cell(row, 8).number_format = "yyyy-mm-dd"
        for column in (5, 6, 9, 14, 15, 17, 18, 19, 20, 21, 25, 27):
            ws.cell(row, column).number_format = CURRENCY_FORMAT
        ws.cell(row, 22).number_format = PERCENT_FORMAT
        for column in (4, 10, 16, 26):
            ws.cell(row, column).number_format = "0"

    widths = {
        1: 14, 2: 14, 3: 14, 4: 15, 5: 12, 6: 16, 7: 28,
        8: 13, 9: 12, 10: 14, 11: 28, 12: 16, 13: 30,
        14: 13, 15: 16, 16: 18, 17: 15, 18: 16, 19: 16,
        20: 18, 21: 18, 22: 18, 23: 24, 24: 14, 25: 14,
        26: 16, 27: 14, 28: 26, 29: 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width

    record_type = DataValidation(
        type="list",
        formula1='"止损计划,每日跟踪"',
        allow_blank=True,
    )
    _add_validation(ws, record_type, f"A2:A{end_row}")
    trade_id = DataValidation(type="list", formula1="=_TradeIds", allow_blank=True)
    _add_validation(ws, trade_id, f"B2:B{end_row}")
    stage_number = DataValidation(
        type="whole",
        operator="greaterThanOrEqual",
        formula1="1",
        allow_blank=True,
    )
    _add_validation(ws, stage_number, f"D2:D{end_row}")
    positive_value = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
    )
    _add_validation(ws, positive_value, f"E2:F{end_row} I2:J{end_row} R2:R{end_row} Y2:Y{end_row}")
    record_date = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2000,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )
    _add_validation(ws, record_date, f"H2:H{end_row}")
    market_stage = DataValidation(
        type="list",
        formula1='"吸筹,启动,趋势上行,加速上涨,高位分歧,回调确认,破位,其他"',
        allow_blank=True,
    )
    _add_validation(ws, market_stage, f"L2:L{end_row}")
    sell_action = DataValidation(
        type="list",
        formula1='"继续持有,执行卖出"',
        allow_blank=True,
    )
    _add_validation(ws, sell_action, f"X2:X{end_row}")
    sell_fee = DataValidation(
        type="decimal",
        operator="greaterThanOrEqual",
        formula1="0",
        allow_blank=True,
    )
    _add_validation(ws, sell_fee, f"AA2:AA{end_row}")
    ws.conditional_formatting.add(
        f"A2:AC{end_row}",
        FormulaRule(formula=['$A2="止损计划"'], fill=YELLOW_FILL),
    )
    ws.conditional_formatting.add(
        f"W2:W{end_row}",
        FormulaRule(formula=['LEFT($W2,4)="触发止损"'], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        f"AB2:AC{end_row}",
        FormulaRule(
            formula=['OR(LEFT($AB2,2)="违规",ISNUMBER(SEARCH("无效",$AC2)))'],
            fill=RED_FILL,
        ),
    )
    _add_table(ws, "PositionTracking", f"A1:AC{end_row}")
    ws.sheet_view.showGridLines = False
    return ws


def _build_technical_indicator_sheet(wb: Workbook, end_row: int = 201):
    ws = wb.create_sheet("技术指标")
    _apply_header(ws, ["技术指标", "分类", "说明"])
    for row, values in enumerate(TECHNICAL_INDICATORS, start=2):
        for column, value in enumerate(values, start=1):
            ws.cell(row, column).value = value
    for row in range(2, end_row + 1):
        for column in range(1, 4):
            cell = ws.cell(row, column)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 56
    _add_table(ws, "TechnicalIndicatorCatalog", f"A1:C{end_row}")
    wb.defined_names.add(
        DefinedName(
            "技术指标列表",
            attr_text=f"'技术指标'!$A$2:$A${end_row}",
        )
    )
    ws["A1"].comment = Comment(
        "可在本列继续添加或修改指标；买入理由页的三个下拉框会引用A2:A201。",
        "Codex",
    )
    return ws


def _style_panel(ws) -> None:
    ws.freeze_panes = "A2"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 72
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _build_statistics_sheet(wb: Workbook):
    ws = wb.create_sheet("多次统计数据")
    rows = [
        ("指标", "当前统计", "口径说明"),
        ("已完成交易数", "=COUNT('单次交易'!M2:M201)", "存在卖出日期的交易"),
        ("盈利交易数", '=COUNTIF(\'单次交易\'!X2:X201,">0")', "实际盈亏金额大于0"),
        ("亏损交易数", '=COUNTIF(\'单次交易\'!X2:X201,"<0")', "实际盈亏金额小于0"),
        (
            "持平交易数",
            '=COUNTIFS(\'单次交易\'!M2:M201,"<>",\'单次交易\'!X2:X201,"=0")',
            "已卖出且实际盈亏为0；不进入胜负样本",
        ),
        ("盈利概率", '=IFERROR(B3/(B3+B4),"")', "盈利数÷胜负交易数"),
        ("亏损概率", '=IFERROR(B4/(B3+B4),"")', "亏损数÷胜负交易数"),
        (
            "平均盈利百分比",
            '=IFERROR(AVERAGEIF(\'单次交易\'!X2:X201,">0",\'单次交易\'!AE2:AE201),"")',
            "盈利交易实际账户收益率的平均值",
        ),
        (
            "平均亏损百分比",
            '=IFERROR(-AVERAGEIF(\'单次交易\'!X2:X201,"<0",\'单次交易\'!AE2:AE201),"")',
            "亏损交易实际账户收益率绝对值的平均值",
        ),
        (
            "盈利的持有总天数",
            '=SUMIF(\'单次交易\'!X2:X201,">0",\'单次交易\'!AA2:AA201)',
            "自然日",
        ),
        (
            "亏损的持有总天数",
            '=SUMIF(\'单次交易\'!X2:X201,"<0",\'单次交易\'!AA2:AA201)',
            "自然日",
        ),
        (
            "平均交易金额",
            '=IFERROR(SUMPRODUCT(\'单次交易\'!E2:E201,\'单次交易\'!H2:H201)/COUNT(\'单次交易\'!H2:H201),"")',
            "买入价×实际买入股数的平均值，包含未卖出交易",
        ),
        ("期望收益率", '=IF(OR(B6="",B7="",B8="",B9=""),"",B6*B8-B7*B9)', "盈利贡献减亏损贡献"),
        (
            "迄今为止的复利净利润率",
            '=IF(COUNT(\'单次交易\'!AE2:AE201)=0,"",'
            'EXP(SUMPRODUCT(IFERROR(LN(1+\'单次交易\'!AE2:AE201),0)))-1)',
            "按每笔实际账户收益率逐笔连乘",
        ),
    ]
    for row in rows:
        ws.append(row)
    _style_panel(ws)
    for row in range(2, 15):
        ws.cell(row, 2).fill = FORMULA_FILL
    for row in (6, 7, 8, 9, 13, 14):
        ws.cell(row, 2).number_format = PERCENT_FORMAT
    for row in (12,):
        ws.cell(row, 2).number_format = CURRENCY_FORMAT
    return ws


def _build_account_sheet(wb: Workbook):
    ws = wb.create_sheet("账户数据")
    rows = [
        ("指标", "值", "录入/计算说明"),
        ("初始总金额", None, "手动输入；当前金额从此值开始累计"),
        (
            "当前总金额",
            "=IF(B2=\"\",\"\",B2+SUM('单次交易'!X2:X201))",
            "初始总金额＋全部已实现盈亏",
        ),
        ("单次交易允许的亏损比例", 0.01, "默认1%，可调整；开仓时复制为交易快照"),
        ("单次交易允许的亏损金额", '=IF(B3="","",B3*B4)', "当前总金额×单次允许亏损比例"),
        ("当月允许最大亏损比例", 0.05, "默认5%，可调整"),
        ("当月允许最大亏损金额", '=IF(B3="","",B3*B6)', "当前总金额×当月最大亏损比例"),
        (
            "当月亏损金额",
            '=SUMIFS(\'单次交易\'!X2:X201,\'单次交易\'!X2:X201,"<0",'
            '\'单次交易\'!M2:M201,">="&EOMONTH(TODAY(),-1)+1,'
            '\'单次交易\'!M2:M201,"<="&EOMONTH(TODAY(),0))',
            "按卖出日期统计本月已实现亏损，保留负号",
        ),
        (
            "当前未平仓理论亏损",
            '=IFERROR(SUMPRODUCT((\'单次交易\'!M2:M201="")*'
            '(\'单次交易\'!H2:H201>0)*'
            '(\'单次交易\'!E2:E201>\'单次交易\'!N2:N201)*'
            '\'单次交易\'!H2:H201*'
            '(\'单次交易\'!E2:E201-\'单次交易\'!N2:N201)),0)',
            "未填写卖出日期的实际持仓×买入价与止损价之差",
        ),
        (
            "当月剩余可开仓风险额度",
            '=IF(OR(B7="",B9=""),"",B7-B9)',
            "当月允许最大亏损金额－当前未平仓理论亏损",
        ),
        (
            "账户实际累计收益率",
            '=IF(OR(B2="",B3="",B2<=0),"",B3/B2-1)',
            "当前总金额÷初始总金额－1；无入金出金时为账户权威累计收益",
        ),
    ]
    for row in rows:
        ws.append(row)
    _style_panel(ws)
    for cell in ("B2", "B4", "B6"):
        ws[cell].fill = INPUT_FILL
    for cell in ("B3", "B5", "B7", "B8", "B9", "B10", "B11"):
        ws[cell].fill = FORMULA_FILL
    for cell in ("B2", "B3", "B5", "B7", "B8", "B9", "B10"):
        ws[cell].number_format = CURRENCY_FORMAT
    for cell in ("B4", "B6", "B11"):
        ws[cell].number_format = PERCENT_FORMAT
    positive_amount = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=False,
    )
    _add_validation(ws, positive_amount, "B2")
    risk_rates = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=False,
    )
    _add_validation(ws, risk_rates, "B4 B6")
    ws.conditional_formatting.add(
        "B8",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
    )
    ws.conditional_formatting.add(
        "B10",
        CellIsRule(operator="lessThan", formula=["0"], fill=RED_FILL),
    )
    return ws


def _build_target_sheet(wb: Workbook):
    ws = wb.create_sheet("目标收益")
    rows = [
        ("指标", "值", "口径说明"),
        ("总金额", "='账户数据'!B3", "来自账户当前总金额"),
        ("目标收益率", 0.10, "手动输入"),
        ("目标收益金额", '=IF(OR(B2="",B3=""),"",B2*B3)', "总金额×目标收益率"),
        ("平均交易金额", "='多次统计数据'!B12", "来自历史交易统计"),
        ("历史期望收益率", "='多次统计数据'!B13", "来自历史胜率与平均盈亏"),
        (
            "所需投资笔数",
            '=IF(OR(B3="",B6="",B3<=0,B6<=0),"暂不可计算",ROUNDUP(LN(1+B3)/LN(1+B6),0))',
            "按目标收益率与历史账户期望收益率复利计算并向上取整",
        ),
    ]
    for row in rows:
        ws.append(row)
    _style_panel(ws)
    ws["B3"].fill = INPUT_FILL
    for cell in ("B2", "B4", "B5", "B6", "B7"):
        ws[cell].fill = FORMULA_FILL
    for cell in ("B2", "B4", "B5"):
        ws[cell].number_format = CURRENCY_FORMAT
    for cell in ("B3", "B6"):
        ws[cell].number_format = PERCENT_FORMAT
    target_rate = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=False,
    )
    _add_validation(ws, target_rate, "B3")
    return ws


def generate_sample_transactions(
    as_of_date: date,
    count: int = 100,
) -> list[dict[str, Any]]:
    """Generate deterministic trades for progressive workbook verification."""
    if count < 0 or count > 200:
        raise ValueError("count must be between 0 and 200")

    stock_codes = [
        "600519",
        "000001",
        "300750",
        "002594",
        "601318",
        "688981",
        "000858",
        "600036",
        "002415",
        "601899",
    ]
    sectors = [
        "消费",
        "银行",
        "新能源",
        "汽车",
        "保险",
        "半导体",
        "食品饮料",
        "金融",
        "电子",
        "资源",
    ]
    indicator_names = [item[0] for item in TECHNICAL_INDICATORS]
    start_date = as_of_date - timedelta(days=count)
    current_balance = 100_000.0
    result: list[dict[str, Any]] = []

    for index in range(1, count + 1):
        cycle = index % 10
        if index % 25 == 0:
            outcome = "candidate"
        elif cycle == 0:
            outcome = "open"
        elif cycle in (1, 2, 3, 4):
            outcome = "win"
        elif cycle in (5, 6, 7, 8):
            outcome = "loss"
        else:
            outcome = "flat"

        buy_price = round(10 + (index % 11) * 1.25, 2)
        stop_distance = 0.75 + (index % 4) * 0.25
        stop_price = round(buy_price - stop_distance, 2)
        risk_rate = 0.01
        suggested_shares = calculate_position_size(
            current_balance,
            risk_rate,
            buy_price,
            stop_price,
        )
        if suggested_shares is None or suggested_shares <= 0:
            raise ValueError("sample inputs must produce a valid position")
        position_fraction = (0.2, 0.4, 0.6, 0.8, 1.0)[(index - 1) % 5]
        actual_buy_shares = max(
            100,
            math.floor(suggested_shares * position_fraction / 100) * 100,
        )
        actual_buy_shares = min(actual_buy_shares, suggested_shares)
        if outcome == "candidate":
            actual_buy_shares = None

        buy_date = start_date + timedelta(days=index)
        sell_date = None
        sell_price = None
        buy_fee = 0 if outcome == "flat" else (5 if index % 3 else 0)
        sell_fee = 0 if outcome == "flat" else (5 if index % 4 else 0)
        if outcome == "win":
            sell_price = round(
                buy_price * (1.12 + (index % 3) * 0.01),
                2,
            )
        elif outcome == "loss":
            sell_price = round(
                min(stop_price, buy_price * (0.95 - (index % 2) * 0.01)),
                2,
            )
        elif outcome == "flat":
            sell_price = buy_price
        if outcome in {"win", "loss", "flat"}:
            sell_date = min(
                buy_date + timedelta(days=1 + index % 7),
                as_of_date,
            )

        indicators = tuple(
            indicator_names[(index - 1 + offset * 5) % len(indicator_names)]
            for offset in range(3)
        )
        score = {
            "win": "4-良好",
            "loss": "2-较差",
            "flat": "3-一般",
            "open": "",
            "candidate": "",
        }[outcome]
        sources = (
            f"{indicators[0]}与{indicators[1]}共同确认",
            f"止损设在{stop_price:.2f}",
            f"目标参考{indicators[2]}",
            "按计划成交" if sell_date is not None else "",
        )
        item = {
            "trade_id": f"T{as_of_date.year}{index:03d}",
            "stock_code": stock_codes[(index - 1) % len(stock_codes)],
            "sector": sectors[(index - 1) % len(sectors)],
            "account_snapshot": round(current_balance, 2),
            "risk_rate": risk_rate,
            "buy_price": buy_price,
            "suggested_shares": suggested_shares,
            "actual_buy_shares": actual_buy_shares,
            "buy_date": buy_date,
            "expected_sell_price": round(buy_price * 1.10, 2),
            "sell_price": sell_price,
            "sell_date": sell_date,
            "stop_price": stop_price,
            "buy_fee": buy_fee,
            "sell_fee": sell_fee,
            "sources": sources,
            "indicators": indicators,
            "summary": f"{outcome}场景，第{index}笔逐步验证数据",
            "score": score,
            "outcome": outcome,
        }
        result.append(item)

        if sell_date is not None and actual_buy_shares is not None:
            current_balance += calculate_realized_pnl(
                buy_price,
                actual_buy_shares,
                sell_price,
                actual_buy_shares,
                buy_fee,
                sell_fee,
            )

    return result


def sample_trade_metrics(as_of_date: date) -> list[dict[str, Any]]:
    """Return the calculation-grain records represented by sample workbook rows."""
    result = []
    for item in generate_sample_transactions(as_of_date):
        shares = item["actual_buy_shares"]
        pnl = None
        return_rate = None
        account_return = None
        hold_days = None
        if (
            shares is not None
            and item["sell_price"] is not None
            and item["sell_date"] is not None
        ):
            pnl = calculate_realized_pnl(
                item["buy_price"],
                shares,
                item["sell_price"],
                shares,
                item["buy_fee"],
                item["sell_fee"],
            )
            return_rate = calculate_return_rate(
                pnl,
                item["buy_price"],
                shares,
                item["buy_fee"],
            )
            account_return = calculate_account_return(
                pnl,
                item["account_snapshot"],
            )
            hold_days = (item["sell_date"] - item["buy_date"]).days
        result.append(
            {
                "pnl": pnl,
                "return_rate": return_rate,
                "account_return": account_return,
                "account_snapshot": item["account_snapshot"],
                "hold_days": hold_days,
                "trade_amount": (
                    item["buy_price"] * shares
                    if shares is not None
                    else None
                ),
                "sell_date": item["sell_date"],
                "buy_price": item["buy_price"],
                "stop_price": item["stop_price"],
                "actual_buy_shares": shares,
            }
        )
    return result


def append_trade_to_workbook(
    wb: Workbook,
    item: Mapping[str, Any],
    row: int,
) -> None:
    """Append one trade input record without touching adjacent rows."""
    trade_ws = wb["单次交易"]
    values = {
        1: item["trade_id"],
        2: item["stock_code"],
        3: item["account_snapshot"],
        4: item["risk_rate"],
        5: item["buy_price"],
        8: item["actual_buy_shares"],
        9: item["buy_date"],
        10: item["expected_sell_price"],
        11: item["sell_price"],
        13: item["sell_date"],
        14: item["stop_price"],
        15: item["buy_fee"],
        16: item["sell_fee"],
        17: item["sources"][0],
        18: item["sources"][1],
        19: item["sources"][2],
        20: item["sources"][3],
        30: item["score"],
    }
    for column, value in values.items():
        trade_ws.cell(row, column).value = value


def append_reason_to_workbook(
    wb: Workbook,
    item: Mapping[str, Any],
    row: int,
) -> None:
    """Append one buy-reason input record without touching adjacent rows."""
    reason_ws = wb["买入理由"]
    values = (
        item["trade_id"],
        item["stock_code"],
        item["sector"],
        "买入",
        item["buy_date"],
        item["indicators"][0],
        item["indicators"][1],
        item["indicators"][2],
        item["summary"],
    )
    for column, value in enumerate(values, start=1):
        reason_ws.cell(row, column).value = value


def _populate_sample_data(wb: Workbook, as_of_date: date) -> None:
    wb["账户数据"]["B2"] = 100_000
    wb["目标收益"]["B3"] = 0.10
    tracking = wb["持仓跟踪"]
    for row, item in enumerate(
        generate_sample_transactions(as_of_date),
        start=2,
    ):
        append_trade_to_workbook(wb, item, row)
        append_reason_to_workbook(wb, item, row)
        tracking.cell(row, 1).value = "止损计划"
        tracking.cell(row, 2).value = item["trade_id"]
        tracking.cell(row, 4).value = 1
        tracking.cell(row, 5).value = item["expected_sell_price"]
        tracking.cell(row, 6).value = item["buy_price"]
        tracking.cell(row, 7).value = "达到首个目标后把止损提高到买入价"
        expectation_type = TRADE_EXPECTATION_TYPES[(row - 2) % 3]
        matching_period = {
            "短期博反弹": "4～10个交易日",
            "突破冲新高": "11～20个交易日",
            "趋势波段": "21～60个交易日",
        }[expectation_type]
        wb["单次交易"].cell(row, 33).value = expectation_type
        wb["单次交易"].cell(row, 34).value = matching_period
        wb["单次交易"].cell(row, 35).value = (
            f"{expectation_type}样本：按预设周期验证交易计划"
        )


def _consecutive_loss_alert_formula(
    row: int,
    require_stop_plan: bool = False,
    require_expectation: bool = False,
) -> str:
    allowed_check = (
        f'IF(\'账户数据\'!$B$23="接近锁仓",'
        '"允许开仓（接近锁仓）","允许开仓")'
    )
    if require_expectation:
        allowed_check = (
            f'IF(AJ{row}="周期与交易预期不匹配，请重新确认",'
            f'IF(\'账户数据\'!$B$23="接近锁仓",'
            '"允许开仓（接近锁仓；周期需复核）","允许开仓（周期需复核）"),'
            f'IF(\'账户数据\'!$B$23="接近锁仓",'
            '"允许开仓（接近锁仓）","允许开仓"))'
        )
    risk_check = (
        f"IF('账户数据'!$B$9+F{row}*(E{row}-N{row})"
        ">='账户数据'!$B$7,"
        '"禁止开仓：风险额度不足",'
        f'{allowed_check})'
    )
    opening_check = risk_check
    if require_stop_plan:
        opening_check = (
            "IF(COUNTIFS('持仓跟踪'!$A$2:$A$501,\"止损计划\","
            f"'持仓跟踪'!$B$2:$B$501,A{row},"
            "'持仓跟踪'!$AC$2:$AC$501,\"计划有效\")=0,"
            f'"禁止开仓：请先填写止损计划",{risk_check})'
        )
    if require_expectation:
        opening_check = (
            f'IF(OR(AG{row}="",AH{row}="",AI{row}=""),'
            f'"禁止开仓：交易预期未填写完整",{opening_check})'
        )
    return (
        f'=IF(A{row}="","",IF(H{row}<>"","已开仓",'
        f'IF(\'账户数据\'!$B$23="已锁仓",'
        '"禁止开仓：连续亏损达到上限",'
        f'IF(\'账户数据\'!$B$23="解锁信息不完整",'
        '"禁止开仓：请完整填写解锁信息",'
        f'IF(OR(F{row}="",E{row}="",N{row}="",E{row}<=N{row},'
        "'账户数据'!$B$7=\"\"),\"\","
        f"{opening_check})))))"
    )


def apply_consecutive_loss_lock(
    wb: Workbook,
    end_row: int = 201,
) -> Workbook:
    """Add formula-only consecutive-loss locking to an existing workbook."""
    account = wb["账户数据"]
    trade = wb["单次交易"]
    controls = {
        13: (
            "连续亏损锁仓比例",
            0.06,
            "从最近盈利或有效手动解锁后的下一笔开始累计；默认6%",
        ),
        14: (
            "最近盈利记录序号",
            '=IFERROR(LOOKUP(2,1/((\'单次交易\'!$M$2:$M$201<>"")*'
            '(\'单次交易\'!$X$2:$X$201>0)),'
            "ROW('单次交易'!$X$2:$X$201)-1),0)",
            "按工作表记录顺序查找最近一笔已平仓盈利交易",
        ),
        15: (
            "最近已完成记录序号",
            '=IFERROR(LOOKUP(2,1/(\'单次交易\'!$M$2:$M$201<>""),'
            "ROW('单次交易'!$M$2:$M$201)-1),0)",
            "按工作表记录顺序查找最后一笔已平仓交易",
        ),
        16: (
            "手动解锁截至记录序号",
            None,
            "锁仓复盘后填写当前最近已完成记录序号；须同时填写解锁原因",
        ),
        17: (
            "手动解锁原因",
            None,
            "例如：已完成复盘，开始新的模拟周期",
        ),
        18: (
            "当前周期起点序号",
            '=MAX(B14,IF(AND(ISNUMBER(B16),B16>=1,B16<=B15,B17<>""),B16,0))',
            "取最近盈利序号与有效手动解锁序号中的较大值",
        ),
        19: (
            "周期起点账户金额",
            '=IF(B2="","",IF(B18=0,B2,B2+'
            "SUM('单次交易'!$X$2:INDEX('单次交易'!$X:$X,B18+1))))",
            "初始金额加上周期起点及以前的全部已实现盈亏",
        ),
        20: (
            "当前周期累计亏损",
            '=IF(B18>=B15,0,-SUMPRODUCT('
            "(ROW('单次交易'!$X$2:$X$201)-1>B18)*"
            "('单次交易'!$M$2:$M$201<>\"\")*"
            "(IFERROR('单次交易'!$X$2:$X$201,0)<0)*"
            "IFERROR('单次交易'!$X$2:$X$201,0)))",
            "仅累计周期起点之后的已平仓亏损；盈利会自动成为新起点",
        ),
        21: (
            "连续亏损比例",
            '=IF(OR(B19="",B19<=0),"",B20/B19)',
            "当前周期累计亏损÷周期起点账户金额",
        ),
        22: (
            "锁仓风险使用率",
            '=IF(OR(B13="",B13<=0,B21=""),"",B21/B13)',
            "达到80%预警，达到100%锁仓",
        ),
        23: (
            "连续亏损锁仓状态",
            '=IF(OR(AND(B16<>"",B17=""),AND(B16="",B17<>""),'
            'AND(B16<>"",OR(NOT(ISNUMBER(B16)),B16<1,B16>B15,B16<>INT(B16)))),'
            '"解锁信息不完整",IF(B22>=1,"已锁仓",'
            'IF(B22>=0.8,"接近锁仓","正常")))',
            "锁仓时禁止新开仓；有效手动解锁后从下一笔重新累计",
        ),
    }
    for row, values in controls.items():
        for column, value in enumerate(values, start=1):
            account.cell(row, column).value = value
    _style_panel(account)
    for cell in ("B13", "B16", "B17"):
        account[cell].fill = INPUT_FILL
    for cell in (
        "B14",
        "B15",
        "B18",
        "B19",
        "B20",
        "B21",
        "B22",
        "B23",
    ):
        account[cell].fill = FORMULA_FILL
    account["B13"].number_format = PERCENT_FORMAT
    account["B19"].number_format = CURRENCY_FORMAT
    account["B20"].number_format = CURRENCY_FORMAT
    account["B21"].number_format = PERCENT_FORMAT
    account["B22"].number_format = PERCENT_FORMAT
    for cell in ("B14", "B15", "B16", "B18"):
        account[cell].number_format = "0"

    lock_rate = DataValidation(
        type="decimal",
        operator="between",
        formula1="0",
        formula2="1",
        allow_blank=False,
    )
    _add_validation(account, lock_rate, "B13")
    unlock_sequence = DataValidation(
        type="custom",
        formula1='=OR(B16="",AND(ISNUMBER(B16),B16=INT(B16),B16>=1,B16<=B15))',
        allow_blank=True,
    )
    unlock_sequence.errorTitle = "解锁序号无效"
    unlock_sequence.error = "请填写1到最近已完成记录序号之间的整数，并同时填写解锁原因"
    unlock_sequence.showErrorMessage = True
    _add_validation(account, unlock_sequence, "B16")
    account.conditional_formatting.add(
        "B23",
        FormulaRule(formula=['$B$23="已锁仓"'], fill=RED_FILL),
    )
    account.conditional_formatting.add(
        "B23",
        FormulaRule(
            formula=['OR($B$23="接近锁仓",$B$23="解锁信息不完整")'],
            fill=YELLOW_FILL,
        ),
    )
    account.conditional_formatting.add(
        "B23",
        FormulaRule(formula=['$B$23="正常"'], fill=GREEN_FILL),
    )

    if "_LockStatus" in wb.defined_names:
        del wb.defined_names["_LockStatus"]
    wb.defined_names.add(
        DefinedName("_LockStatus", attr_text="'账户数据'!$B$23")
    )
    for row in range(2, end_row + 1):
        trade.cell(row, 7).value = _consecutive_loss_alert_formula(row)

    trade.data_validations.dataValidation = [
        validation
        for validation in trade.data_validations.dataValidation
        if f"H2:H{end_row}" not in str(validation.sqref)
    ]
    actual_shares = DataValidation(
        type="custom",
        formula1=(
            '=OR(H2="",AND(_LockStatus<>"已锁仓",'
            '_LockStatus<>"解锁信息不完整",ISNUMBER(H2),'
            'H2>0,MOD(H2,100)=0))'
        ),
        allow_blank=True,
    )
    actual_shares.errorTitle = "当前禁止开仓或股数无效"
    actual_shares.error = (
        "请先完成解锁，并确保实际买入股数为大于0的100股整数倍"
    )
    actual_shares.showErrorMessage = True
    _add_validation(trade, actual_shares, f"H2:H{end_row}")
    trade.conditional_formatting.add(
        f"G2:G{end_row}",
        FormulaRule(
            formula=['LEFT($G2,4)="禁止开仓"'],
            fill=RED_FILL,
        ),
    )
    trade.conditional_formatting.add(
        f"G2:G{end_row}",
        FormulaRule(
            formula=['$G2="允许开仓（接近锁仓）"'],
            fill=YELLOW_FILL,
        ),
    )
    trade.conditional_formatting.add(
        f"G2:G{end_row}",
        FormulaRule(formula=['$G2="允许开仓"'], fill=GREEN_FILL),
    )
    account.sheet_view.showGridLines = False
    wb.active = wb.sheetnames.index("账户数据")
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    return wb


def upgrade_workbook_with_consecutive_loss_lock(
    source: str | Path,
    destination: str | Path,
) -> Path:
    """Preserve workbook data while adding the formula-only lock controls."""
    source_path = Path(source)
    destination_path = Path(destination)
    workbook = load_workbook(source_path, data_only=False)
    apply_consecutive_loss_lock(workbook)
    workbook.save(destination_path)
    return destination_path


def _tracking_sell_lookup_formula(
    trade_row: int,
    tracking_column: str,
) -> str:
    return (
        f'=IF(A{trade_row}="","",IFERROR(LOOKUP(2,1/'
        f"(('持仓跟踪'!$B$2:$B$501=A{trade_row})*"
        "('持仓跟踪'!$X$2:$X$501=\"执行卖出\")*"
        f"('持仓跟踪'!${tracking_column}$2:${tracking_column}$501<>\"\")),"
        f"'持仓跟踪'!${tracking_column}$2:${tracking_column}$501),\"\"))"
    )


def apply_dynamic_stop_tracking(
    wb: Workbook,
    tracking_end_row: int = 501,
    trade_end_row: int = 201,
) -> Workbook:
    """Add the combined stop-plan and daily position-tracking workflow."""
    if "持仓跟踪" in wb.sheetnames:
        raise ValueError("持仓跟踪工作表已存在；为避免覆盖数据，不重复创建")
    tracking = _build_tracking_sheet(wb, tracking_end_row)
    trade = wb["单次交易"]

    for name, reference in {
        "_TradeIds": "'单次交易'!$A$2:$A$201",
        "_TrackType": "'持仓跟踪'!$A$2:$A$501",
        "_TrackTradeId": "'持仓跟踪'!$B$2:$B$501",
        "_TrackRule": "'持仓跟踪'!$AC$2:$AC$501",
    }.items():
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names.add(DefinedName(name, attr_text=reference))

    for row in range(2, trade_end_row + 1):
        trade.cell(row, 7).value = _consecutive_loss_alert_formula(
            row,
            require_stop_plan=True,
        )
    trade.data_validations.dataValidation = [
        validation
        for validation in trade.data_validations.dataValidation
        if f"H2:H{trade_end_row}" not in str(validation.sqref)
    ]
    actual_shares = DataValidation(
        type="custom",
        formula1=(
            '=OR(H2="",AND(_LockStatus<>"已锁仓",'
            '_LockStatus<>"解锁信息不完整",'
            'COUNTIFS(_TrackType,"止损计划",_TrackTradeId,A2,'
            '_TrackRule,"计划有效")>0,ISNUMBER(H2),H2>0,'
            'MOD(H2,100)=0))'
        ),
        allow_blank=True,
    )
    actual_shares.errorTitle = "当前禁止开仓、止损计划缺失或股数无效"
    actual_shares.error = (
        "请先完成有效止损计划及解锁，并确保股数为大于0的100股整数倍"
    )
    actual_shares.showErrorMessage = True
    _add_validation(trade, actual_shares, f"H2:H{trade_end_row}")

    existing_rows = [
        row
        for row in range(2, trade_end_row + 1)
        if trade.cell(row, 1).value not in (None, "")
    ]
    first_future_row = max(existing_rows, default=1) + 1
    for row in range(first_future_row, trade_end_row + 1):
        trade.cell(row, 11).value = _tracking_sell_lookup_formula(row, "Y")
        trade.cell(row, 12).value = _tracking_sell_lookup_formula(row, "Z")
        trade.cell(row, 13).value = _tracking_sell_lookup_formula(row, "H")
        trade.cell(row, 16).value = _tracking_sell_lookup_formula(row, "AA")
        for column in (11, 12, 13, 16):
            trade.cell(row, column).fill = FORMULA_FILL

    trade["G1"].comment = Comment(
        "开仓前必须在“持仓跟踪”填写至少一条规则有效的止损计划；"
        "同时检查连续亏损锁仓和原有风险额度。",
        "Codex",
    )
    tracking["A1"].comment = Comment(
        "同一张表同时记录止损计划和每日跟踪；先选择记录类型，再填写对应的蓝色输入列。",
        "Codex",
    )
    tracking["S1"].comment = Comment(
        "当前有效止损取初始止损、已激活计划止损和历次人工上调止损的最大值，只升不降。",
        "Codex",
    )
    wb.active = wb.sheetnames.index("持仓跟踪")
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    return wb


def upgrade_workbook_with_dynamic_stop(
    source: str | Path,
    destination: str | Path,
) -> Path:
    """Preserve history and add the dynamic-stop tracking workflow."""
    source_path = Path(source)
    destination_path = Path(destination)
    workbook = load_workbook(source_path, data_only=False)
    apply_dynamic_stop_tracking(workbook)
    workbook.save(destination_path)
    return destination_path


def _expectation_match_formula(row: int) -> str:
    return (
        f'=IF(OR(AG{row}="",AH{row}=""),"",IF(OR('
        f'AND(AG{row}="短期博反弹",OR(AH{row}="1～3个交易日",'
        f'AH{row}="4～10个交易日")),'
        f'AND(AG{row}="突破冲新高",OR(AH{row}="4～10个交易日",'
        f'AH{row}="11～20个交易日")),'
        f'AND(AG{row}="趋势波段",OR(AH{row}="11～20个交易日",'
        f'AH{row}="21～60个交易日",AH{row}="60个交易日以上"))),'
        '"匹配","周期与交易预期不匹配，请重新确认"))'
    )


def apply_trade_expectation_fields(
    wb: Workbook,
    trade_end_row: int = 201,
) -> Workbook:
    """Require a trade thesis, holding range, and written rationale."""
    trade = wb["单次交易"]
    if trade["AG1"].value not in (None, ""):
        raise ValueError("交易预期字段已存在；为避免覆盖数据，不重复创建")

    if trade["AF1"].value in (None, ""):
        trade["AF1"] = "备注"
        trade["AF1"].fill = HEADER_FILL
        trade["AF1"].font = HEADER_FONT
        trade["AF1"].alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        trade["AF1"].border = THIN_BORDER
        for row in range(2, trade_end_row + 1):
            trade.cell(row, 32).fill = INPUT_FILL
            trade.cell(row, 32).border = THIN_BORDER
            trade.cell(row, 32).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

    expectation_headers = (
        "交易预期类型",
        "预期持有周期",
        "预期选择理由",
        "周期匹配检查",
    )
    for column, header in enumerate(expectation_headers, start=33):
        cell = trade.cell(1, column)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = THIN_BORDER
    for row in range(2, trade_end_row + 1):
        for column in (33, 34, 35):
            trade.cell(row, column).fill = INPUT_FILL
            trade.cell(row, column).border = THIN_BORDER
            trade.cell(row, column).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
        match_cell = trade.cell(row, 36)
        match_cell.value = _expectation_match_formula(row)
        match_cell.fill = FORMULA_FILL
        match_cell.border = THIN_BORDER
        match_cell.alignment = Alignment(vertical="top", wrap_text=True)

    trade.column_dimensions["AF"].width = max(
        trade.column_dimensions["AF"].width or 0,
        20,
    )
    trade.column_dimensions["AG"].width = 18
    trade.column_dimensions["AH"].width = 20
    trade.column_dimensions["AI"].width = 42
    trade.column_dimensions["AJ"].width = 32
    expectation_type = DataValidation(
        type="list",
        formula1='"短期博反弹,突破冲新高,趋势波段"',
        allow_blank=True,
    )
    expectation_type.errorTitle = "交易预期无效"
    expectation_type.error = "请从短期博反弹、突破冲新高、趋势波段中选择"
    expectation_type.showErrorMessage = True
    _add_validation(trade, expectation_type, f"AG2:AG{trade_end_row}")
    holding_period = DataValidation(
        type="list",
        formula1=(
            '"1～3个交易日,4～10个交易日,11～20个交易日,'
            '21～60个交易日,60个交易日以上"'
        ),
        allow_blank=True,
    )
    holding_period.errorTitle = "预期持有周期无效"
    holding_period.error = "请从预设的交易日范围中选择"
    holding_period.showErrorMessage = True
    _add_validation(trade, holding_period, f"AH2:AH{trade_end_row}")
    trade.conditional_formatting.add(
        f"AJ2:AJ{trade_end_row}",
        FormulaRule(
            formula=['$AJ2="周期与交易预期不匹配，请重新确认"'],
            fill=YELLOW_FILL,
        ),
    )
    trade.conditional_formatting.add(
        f"AJ2:AJ{trade_end_row}",
        FormulaRule(formula=['$AJ2="匹配"'], fill=GREEN_FILL),
    )

    for row in range(2, trade_end_row + 1):
        trade.cell(row, 7).value = _consecutive_loss_alert_formula(
            row,
            require_stop_plan=True,
            require_expectation=True,
        )
    trade.data_validations.dataValidation = [
        validation
        for validation in trade.data_validations.dataValidation
        if f"H2:H{trade_end_row}" not in str(validation.sqref)
    ]
    actual_shares = DataValidation(
        type="custom",
        formula1=(
            '=OR(H2="",AND(_LockStatus<>"已锁仓",'
            '_LockStatus<>"解锁信息不完整",AG2<>"",AH2<>"",AI2<>"",'
            'COUNTIFS(_TrackType,"止损计划",_TrackTradeId,A2,'
            '_TrackRule,"计划有效")>0,ISNUMBER(H2),H2>0,'
            'MOD(H2,100)=0))'
        ),
        allow_blank=True,
    )
    actual_shares.errorTitle = "交易计划不完整、当前禁止开仓或股数无效"
    actual_shares.error = (
        "请填写交易预期、持有周期、选择理由和有效止损计划，"
        "完成解锁，并确保股数为大于0的100股整数倍"
    )
    actual_shares.showErrorMessage = True
    _add_validation(trade, actual_shares, f"H2:H{trade_end_row}")

    table = next(iter(trade.tables.values()))
    existing_column_count = len(table.tableColumns)
    for column in range(existing_column_count + 1, 37):
        table.tableColumns.append(
            TableColumn(id=column, name=str(trade.cell(1, column).value))
        )
    table.ref = f"A1:AJ{trade_end_row}"
    if table.autoFilter is not None:
        table.autoFilter.ref = table.ref
    trade["AG1"].comment = Comment(
        "开仓前必选：短期博反弹、突破冲新高或趋势波段。"
        "同时必须选择预期持有周期并填写选择理由。",
        "Codex",
    )
    wb.active = wb.sheetnames.index("单次交易")
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    return wb


def upgrade_workbook_with_trade_expectations(
    source: str | Path,
    destination: str | Path,
) -> Path:
    """Preserve V3 data and add mandatory trade-expectation fields."""
    source_path = Path(source)
    destination_path = Path(destination)
    workbook = load_workbook(source_path, data_only=False)
    apply_trade_expectation_fields(workbook)
    workbook.save(destination_path)
    return destination_path


def build_workbook(
    with_sample_data: bool = False,
    as_of_date: date | None = None,
) -> Workbook:
    """Build the six-sheet trading workbook."""
    wb = Workbook()
    wb.remove(wb.active)
    _build_trade_sheet(wb)
    _build_reason_sheet(wb)
    _build_statistics_sheet(wb)
    _build_account_sheet(wb)
    _build_target_sheet(wb)
    _build_technical_indicator_sheet(wb)
    apply_consecutive_loss_lock(wb)
    apply_dynamic_stop_tracking(wb)
    apply_trade_expectation_fields(wb)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    if with_sample_data:
        _populate_sample_data(wb, as_of_date or date.today())
    return wb


def write_workbooks(
    output_dir: str | Path,
    as_of_date: date | None = None,
) -> tuple[Path, Path]:
    """Write the clean template and synthetic test workbook."""
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    effective_date = as_of_date or date.today()
    clean_path = destination / "交易管理系统.xlsx"
    sample_path = destination / "交易管理系统_测试版.xlsx"
    build_workbook(with_sample_data=False).save(clean_path)
    build_workbook(
        with_sample_data=True,
        as_of_date=effective_date,
    ).save(sample_path)
    return clean_path, sample_path


def main(
    output_dir: str | Path = ".",
    as_of_date: date | None = None,
) -> tuple[Path, Path]:
    """Generate both user-facing workbook deliverables."""
    paths = write_workbooks(output_dir, as_of_date=as_of_date)
    for path in paths:
        print(path)
    return paths


if __name__ == "__main__":
    main()
