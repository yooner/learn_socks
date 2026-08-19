import math
import importlib
import shutil
import subprocess
import tempfile
import unittest
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
import trading_workbook as tw
from trading_workbook import (
    calculate_account_return,
    calculate_compound_return,
    calculate_compound_loss_ceiling,
    calculate_monthly_loss,
    calculate_position_size,
    calculate_realized_pnl,
    calculate_required_trades,
    calculate_return_rate,
    summarize_trades,
)

EXPECTED_SHEETS = [
    "单次交易",
    "持仓跟踪",
    "买入理由",
    "多次统计数据",
    "账户数据",
    "目标收益",
    "技术指标",
]

TRADE_HEADERS = [
    "交易编号",
    "股票代码",
    "买入时账户金额",
    "本次允许亏损比例",
    "第一批买入价",
    "买入建议股数",
    "开仓风险告警",
    "第一批买入股数",
    "首次买入日期",
    "期望卖出价",
    "实际卖出价",
    "卖出股数",
    "卖出日期",
    "止损价",
    "买入总费用",
    "卖出费用",
    "买入价的由来",
    "止损价的由来",
    "期望卖出价的由来",
    "实际卖出价的由来",
    "期望盈利比例",
    "期望止损比例",
    "盈亏比",
    "实际盈亏金额",
    "单笔仓位收益率",
    "账户收益率与平均盈利率差值",
    "持有天数",
    "复利容许平均亏损上限",
    "复利风险判断",
    "交易打分评价",
    "实际账户收益率",
]


class CalculationTests(unittest.TestCase):
    def test_tranche_position_calculates_weighted_cost_shares_and_risk(self):
        result = tw.calculate_tranche_position(
            [(10.0, 1_000), (8.0, 500), (12.0, 500)],
            effective_stop=9.0,
        )

        self.assertEqual(result["total_shares"], 2_000)
        self.assertAlmostEqual(result["buy_amount"], 20_000.0)
        self.assertAlmostEqual(result["weighted_buy_price"], 10.0)
        self.assertAlmostEqual(result["current_risk"], 2_500.0)

    def test_tranche_position_never_counts_negative_risk(self):
        result = tw.calculate_tranche_position(
            [(10.0, 100), (8.0, 100), (None, None)],
            effective_stop=9.0,
        )

        self.assertEqual(result["current_risk"], 100.0)

    def test_closed_tranche_position_has_zero_current_risk(self):
        result = tw.calculate_tranche_position(
            [(10.0, 100), (8.0, 100), (None, None)],
            effective_stop=7.0,
            is_closed=True,
        )

        self.assertEqual(result["current_risk"], 0.0)

    def test_tranche_rules_enforce_pair_order_lots_lock_and_risk(self):
        cases = [
            (
                None,
                [(None, None)] * 3,
                False,
                0,
                1_000,
                0,
                5_000,
                "",
            ),
            (
                "T1",
                [(10, None), (None, None), (None, None)],
                False,
                0,
                1_000,
                0,
                5_000,
                "违规：第一批价格或股数缺失",
            ),
            (
                "T1",
                [(10, 100), (9, None), (None, None)],
                False,
                0,
                1_000,
                0,
                5_000,
                "违规：第二批价格与股数须成对填写",
            ),
            (
                "T1",
                [(10, 100), (9, 100), (8, None)],
                False,
                0,
                1_000,
                0,
                5_000,
                "违规：第三批价格与股数须成对填写",
            ),
            (
                "T1",
                [(10, 100), (None, None), (8, 100)],
                False,
                0,
                1_000,
                0,
                5_000,
                "违规：必须先完成第二批",
            ),
            (
                "T1",
                [(10, 150), (None, None), (None, None)],
                False,
                0,
                1_000,
                0,
                5_000,
                "违规：股数须为100股整数倍",
            ),
            (
                "T1",
                [(10, 100), (9, 100), (None, None)],
                True,
                900,
                1_000,
                900,
                5_000,
                "违规：锁仓期间禁止加仓",
            ),
            (
                "T1",
                [(10, 100), (9, 100), (None, None)],
                False,
                1_001,
                1_000,
                1_001,
                5_000,
                "违规：超过单笔风险上限",
            ),
            (
                "T1",
                [(10, 100), (9, 100), (None, None)],
                False,
                900,
                1_000,
                1_200,
                1_199,
                "违规：超过账户风险上限",
            ),
            (
                "T1",
                [(10, 100), (9, 100), (None, None)],
                False,
                1_000,
                1_000,
                5_000,
                5_000,
                "通过",
            ),
        ]

        for (
            trade_id,
            tranches,
            locked,
            position_risk,
            one_limit,
            total_risk,
            account_limit,
            expected,
        ) in cases:
            with self.subTest(expected=expected):
                self.assertEqual(
                    tw.check_tranche_rules(
                        trade_id,
                        tranches,
                        locked,
                        position_risk,
                        one_limit,
                        total_risk,
                        account_limit,
                    ),
                    expected,
                )

    def test_trade_expectation_period_matching_uses_the_agreed_ranges(self):
        matching_cases = [
            ("短期博反弹", "1～3个交易日"),
            ("短期博反弹", "4～10个交易日"),
            ("突破冲新高", "4～10个交易日"),
            ("突破冲新高", "11～20个交易日"),
            ("趋势波段", "11～20个交易日"),
            ("趋势波段", "21～60个交易日"),
            ("趋势波段", "60个交易日以上"),
        ]
        for trade_type, holding_period in matching_cases:
            with self.subTest(
                trade_type=trade_type,
                holding_period=holding_period,
            ):
                self.assertEqual(
                    tw.check_trade_expectation_period(
                        trade_type,
                        holding_period,
                    ),
                    "匹配",
                )
        self.assertEqual(
            tw.check_trade_expectation_period(
                "短期博反弹",
                "21～60个交易日",
            ),
            "周期与交易预期不匹配，请重新确认",
        )
        self.assertIsNone(tw.check_trade_expectation_period("", "1～3个交易日"))

    def test_dynamic_stop_uses_historical_high_and_never_moves_down(self):
        result = tw.calculate_dynamic_stop_history(
            initial_stop=9,
            plan_levels=[
                {"stage": 1, "activation_price": 11, "stop_price": 10},
                {"stage": 2, "activation_price": 12, "stop_price": 11},
                {"stage": 3, "activation_price": 14, "stop_price": 12.5},
            ],
            closing_prices=[10, 12.2, 10.8],
        )

        self.assertEqual(
            [item["effective_stop"] for item in result],
            [9, 11, 11],
        )
        self.assertEqual(
            [item["activated_stage"] for item in result],
            [0, 2, 2],
        )
        self.assertEqual(result[-1]["stop_status"], "触发止损：应全部卖出")

    def test_manual_dynamic_stop_can_raise_but_cannot_lower_the_stop(self):
        result = tw.calculate_dynamic_stop_history(
            initial_stop=9,
            plan_levels=[
                {"stage": 1, "activation_price": 11, "stop_price": 10},
            ],
            closing_prices=[10, 11.2, 11.1],
            manual_stops=[None, 10.8, 10.2],
        )

        self.assertEqual(
            [item["effective_stop"] for item in result],
            [9, 10.8, 10.8],
        )
        self.assertTrue(result[-1]["manual_stop_rejected"])

    def test_closing_at_the_effective_stop_triggers_a_full_exit(self):
        result = tw.calculate_dynamic_stop_history(
            initial_stop=9,
            plan_levels=[],
            closing_prices=[9],
        )

        self.assertEqual(result[0]["stop_status"], "触发止损：应全部卖出")

    def test_consecutive_loss_cycle_locks_at_six_percent_after_latest_win(self):
        result = tw.calculate_consecutive_loss_lock(
            [100, -30, -40],
            initial_balance=1_000,
            loss_limit_rate=0.06,
        )

        self.assertEqual(result["latest_win_sequence"], 1)
        self.assertEqual(result["cycle_start_sequence"], 1)
        self.assertEqual(result["cycle_start_balance"], 1_100)
        self.assertEqual(result["cycle_loss"], 70)
        self.assertEqual(result["status"], "已锁仓")

    def test_consecutive_loss_cycle_resets_after_every_profit(self):
        result = tw.calculate_consecutive_loss_lock(
            [100, -100, 10, -40],
            initial_balance=1_000,
            loss_limit_rate=0.06,
        )

        self.assertEqual(result["latest_win_sequence"], 3)
        self.assertEqual(result["cycle_start_balance"], 1_010)
        self.assertEqual(result["cycle_loss"], 40)
        self.assertEqual(result["status"], "正常")

    def test_manual_unlock_starts_a_new_cycle_after_the_selected_record(self):
        locked = tw.calculate_consecutive_loss_lock(
            [100, -70],
            initial_balance=1_000,
            loss_limit_rate=0.06,
        )
        incomplete = tw.calculate_consecutive_loss_lock(
            [100, -70],
            initial_balance=1_000,
            loss_limit_rate=0.06,
            manual_unlock_through=2,
        )
        unlocked = tw.calculate_consecutive_loss_lock(
            [100, -70],
            initial_balance=1_000,
            loss_limit_rate=0.06,
            manual_unlock_through=2,
            unlock_reason="模拟复盘完成",
        )

        self.assertEqual(locked["status"], "已锁仓")
        self.assertEqual(incomplete["status"], "解锁信息不完整")
        self.assertEqual(unlocked["cycle_start_sequence"], 2)
        self.assertEqual(unlocked["cycle_loss"], 0)
        self.assertEqual(unlocked["status"], "正常")

    def test_position_size_rounds_down_to_a_share_lot(self):
        self.assertEqual(
            calculate_position_size(
                account_snapshot=100_000,
                risk_rate=0.01,
                buy_price=10,
                stop_price=9,
            ),
            1_000,
        )
        self.assertEqual(
            calculate_position_size(
                account_snapshot=99_780,
                risk_rate=0.01,
                buy_price=25,
                stop_price=23.5,
            ),
            600,
        )

    def test_position_size_rejects_invalid_risk_inputs(self):
        invalid_cases = (
            (0, 0.01, 10, 9),
            (100_000, 0, 10, 9),
            (100_000, 0.01, 10, 10),
            (100_000, 0.01, 10, 11),
        )
        for account, risk_rate, buy_price, stop_price in invalid_cases:
            with self.subTest(
                account=account,
                risk_rate=risk_rate,
                buy_price=buy_price,
                stop_price=stop_price,
            ):
                self.assertIsNone(
                    calculate_position_size(
                        account,
                        risk_rate,
                        buy_price,
                        stop_price,
                    )
                )

    def test_realized_pnl_includes_buy_and_sell_fees(self):
        pnl = calculate_realized_pnl(
            buy_price=10,
            buy_shares=1_000,
            sell_price=11,
            sell_shares=1_000,
            buy_fee=5,
            sell_fee=5,
        )
        self.assertEqual(pnl, 990)
        self.assertAlmostEqual(
            calculate_return_rate(pnl, 10, 1_000, 5),
            990 / 10_005,
        )

    def test_account_return_uses_pre_trade_total_equity(self):
        self.assertEqual(calculate_account_return(1_000, 100_000), 0.01)
        self.assertAlmostEqual(
            calculate_account_return(-2_500, 101_000),
            -2_500 / 101_000,
        )
        self.assertIsNone(calculate_account_return(100, 0))
        self.assertIsNone(calculate_account_return(100, -1))

    def test_compound_return_multiplies_actual_account_returns(self):
        account_returns = [0.01, -2_500 / 101_000]
        self.assertAlmostEqual(
            calculate_compound_return(account_returns),
            -0.015,
        )
        self.assertIsNone(calculate_compound_return([None]))

    def test_summary_uses_only_closed_non_flat_trades_for_win_rate(self):
        trades = [
            {
                "pnl": 990,
                "return_rate": 990 / 10_005,
                "account_return": 990 / 100_000,
                "hold_days": 5,
                "trade_amount": 10_000,
                "sell_date": date(2026, 5, 10),
            },
            {
                "pnl": -1_210,
                "return_rate": -1_210 / 20_005,
                "account_return": -1_210 / 100_990,
                "hold_days": 3,
                "trade_amount": 20_000,
                "sell_date": date(2026, 6, 4),
            },
            {
                "pnl": 1_790,
                "return_rate": 1_790 / 15_005,
                "account_return": 1_790 / 99_780,
                "hold_days": 8,
                "trade_amount": 15_000,
                "sell_date": date(2026, 6, 18),
            },
            {
                "pnl": -732,
                "return_rate": -732 / 18_006,
                "account_return": -732 / 101_570,
                "hold_days": 2,
                "trade_amount": 18_000,
                "sell_date": date(2026, 7, 3),
            },
            {
                "pnl": 0,
                "return_rate": 0,
                "account_return": 0,
                "hold_days": 1,
                "trade_amount": 18_000,
                "sell_date": date(2026, 7, 6),
            },
            {
                "pnl": None,
                "return_rate": None,
                "account_return": None,
                "hold_days": None,
                "trade_amount": 20_000,
                "sell_date": None,
            },
        ]

        result = summarize_trades(trades)

        self.assertEqual(result["completed_count"], 5)
        self.assertEqual(result["win_count"], 2)
        self.assertEqual(result["loss_count"], 2)
        self.assertEqual(result["flat_count"], 1)
        self.assertEqual(result["win_hold_days"], 13)
        self.assertEqual(result["loss_hold_days"], 5)
        self.assertEqual(result["win_rate"], 0.5)
        self.assertEqual(result["loss_rate"], 0.5)
        self.assertAlmostEqual(
            result["average_win"],
            ((990 / 100_000) + (1_790 / 99_780)) / 2,
        )
        self.assertAlmostEqual(
            result["average_loss"],
            ((1_210 / 100_990) + (732 / 101_570)) / 2,
        )
        self.assertGreater(result["average_loss"], 0)
        expected_expectancy = (
            result["win_rate"] * result["average_win"]
            - result["loss_rate"] * result["average_loss"]
        )
        self.assertAlmostEqual(result["expectancy"], expected_expectancy)
        expected_compound = calculate_compound_return(
            [
                990 / 100_000,
                -1_210 / 100_990,
                1_790 / 99_780,
                -732 / 101_570,
                0,
            ]
        )
        self.assertAlmostEqual(result["compound_return"], expected_compound)
        self.assertAlmostEqual(
            result["average_trade_amount"],
            (10_000 + 20_000 + 15_000 + 18_000 + 18_000 + 20_000) / 6,
        )

    def test_compound_loss_ceiling_handles_boundaries(self):
        self.assertIsNone(calculate_compound_loss_ceiling(None, 0.5))
        self.assertIsNone(calculate_compound_loss_ceiling(0.1, 0))
        self.assertIsNone(calculate_compound_loss_ceiling(0.1, 1))
        expected = 1 - (1 + 0.1) ** (-0.5 / (1 - 0.5))
        self.assertAlmostEqual(
            calculate_compound_loss_ceiling(0.1, 0.5),
            expected,
        )

    def test_monthly_loss_preserves_negative_sign(self):
        trades = [
            {"pnl": -732, "sell_date": date(2026, 7, 3)},
            {"pnl": 400, "sell_date": date(2026, 7, 10)},
            {"pnl": -100, "sell_date": date(2026, 6, 30)},
            {"pnl": None, "sell_date": None},
        ]
        self.assertEqual(
            calculate_monthly_loss(trades, as_of_date=date(2026, 7, 24)),
            -732,
        )

    def test_required_trades_rounds_up_and_rejects_non_positive_inputs(self):
        self.assertEqual(calculate_required_trades(1_001, 10_000, 0.025), 5)
        self.assertIsNone(calculate_required_trades(1_000, 0, 0.025))
        self.assertIsNone(calculate_required_trades(1_000, 10_000, 0))
        self.assertIsNone(calculate_required_trades(1_000, 10_000, -0.01))

    def test_sample_generator_produces_100_diverse_sequential_trades(self):
        items = tw.generate_sample_transactions(
            as_of_date=date(2026, 7, 24),
            count=100,
        )
        self.assertEqual(len(items), 100)
        self.assertEqual(len({item["trade_id"] for item in items}), 100)
        self.assertTrue(any(item["sell_date"] is None for item in items))
        self.assertTrue(any(item["actual_buy_shares"] is None for item in items))
        self.assertTrue(
            any(
                item["actual_buy_shares"] is not None
                and item["actual_buy_shares"] < item["suggested_shares"]
                for item in items
            )
        )
        outcomes = {item["outcome"] for item in items}
        self.assertEqual(outcomes, {"win", "loss", "flat", "open", "candidate"})
        indicators = {
            indicator
            for item in items
            for indicator in item["indicators"]
        }
        self.assertIn("蜡烛图", indicators)
        self.assertIn("趋势线", indicators)
        self.assertIn("MACD", indicators)
        sample_summary = summarize_trades(
            tw.sample_trade_metrics(date(2026, 7, 24))
        )
        self.assertGreater(sample_summary["expectancy"], 0)

    def test_sample_metrics_expose_cumulative_position_values(self):
        items = tw.generate_sample_transactions(
            as_of_date=date(2026, 7, 24),
            count=100,
        )
        metrics = tw.sample_trade_metrics(date(2026, 7, 24))
        open_index = next(
            index
            for index, item in enumerate(items)
            if item["actual_buy_shares"] is not None
            and item["sell_date"] is None
        )
        item = items[open_index]
        metric = metrics[open_index]

        expected_amount = item["buy_price"] * item["actual_buy_shares"]
        expected_risk = max(
            item["buy_price"] - item["stop_price"],
            0,
        ) * item["actual_buy_shares"]
        self.assertEqual(metric["buy_amount"], expected_amount)
        self.assertEqual(metric["weighted_buy_price"], item["buy_price"])
        self.assertEqual(
            metric["total_shares"],
            item["actual_buy_shares"],
        )
        self.assertEqual(metric["current_risk"], expected_risk)

    def test_progressive_metric_record_uses_cumulative_position_contract(self):
        validator = importlib.import_module("progressive_workbook_validation")
        item = {
            "buy_price": 10.0,
            "actual_buy_shares": 500,
            "stop_price": 9.0,
            "sell_price": None,
            "sell_date": None,
            "buy_fee": 5.0,
            "sell_fee": 0.0,
            "account_snapshot": 100_000.0,
            "buy_date": date(2026, 7, 1),
        }

        metric = validator._metric_record(item)

        self.assertEqual(metric["buy_amount"], 5_000.0)
        self.assertEqual(metric["weighted_buy_price"], 10.0)
        self.assertEqual(metric["total_shares"], 500)
        self.assertEqual(metric["current_risk"], 500.0)

    def test_append_trade_writes_only_one_requested_row(self):
        workbook = tw.build_workbook(with_sample_data=False)
        item = tw.generate_sample_transactions(
            as_of_date=date(2026, 7, 24),
            count=1,
        )[0]

        tw.append_trade_to_workbook(workbook, item, row=2)
        tw.append_reason_to_workbook(workbook, item, row=2)

        trade = workbook["单次交易"]
        reasons = workbook["买入理由"]
        self.assertEqual(trade["A2"].value, item["trade_id"])
        self.assertEqual(trade["H2"].value, item["actual_buy_shares"])
        self.assertIsNone(trade["A3"].value)
        self.assertEqual(reasons["A2"].value, item["trade_id"])
        self.assertEqual(reasons["F2"].value, item["indicators"][0])
        self.assertIsNone(reasons["A3"].value)

    def test_open_risk_uses_only_unsold_actual_shares(self):
        trades = [
            {
                "buy_price": 10,
                "stop_price": 9,
                "actual_buy_shares": 500,
                "sell_date": None,
            },
            {
                "buy_price": 20,
                "stop_price": 18,
                "actual_buy_shares": 200,
                "sell_date": date(2026, 7, 20),
            },
            {
                "buy_price": 30,
                "stop_price": 31,
                "actual_buy_shares": 300,
                "sell_date": None,
            },
        ]
        self.assertEqual(tw.calculate_open_theoretical_loss(trades), 500)

    def test_risk_status_includes_candidate_at_inclusive_limit(self):
        self.assertEqual(
            tw.calculate_opening_risk_status(4_100, 900, 5_000),
            "禁止开仓",
        )
        self.assertEqual(
            tw.calculate_opening_risk_status(4_000, 900, 5_000),
            "允许开仓",
        )
        self.assertEqual(
            tw.calculate_opening_risk_status(5_100, 0, 5_000),
            "禁止开仓",
        )
        self.assertIsNone(
            tw.calculate_opening_risk_status(0, 0, None),
        )


class WorkbookStructureTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.workbook = tw.build_workbook(with_sample_data=False)

    def test_workbook_has_the_six_agreed_sheets(self):
        self.assertEqual(self.workbook.sheetnames, EXPECTED_SHEETS)

    def test_tracking_sheet_combines_stop_plans_and_daily_records(self):
        tracking = self.workbook["持仓跟踪"]
        expected_headers = [
            "记录类型",
            "交易编号",
            "股票代码",
            "止损阶段序号",
            "激活价",
            "激活后止损价",
            "止损依据",
            "记录日期",
            "收盘价",
            "成交量",
            "量价分析",
            "当前阶段",
            "阶段判断理由",
            "初始止损",
            "历史最高收盘价",
            "已激活止损阶段",
            "当前计划止损",
            "人工上调止损",
            "当前有效止损",
            "下一阶段激活价",
            "下一阶段止损价",
            "距离下一阶段涨幅",
            "止损判断",
            "卖出动作",
            "实际卖出价",
            "实际卖出股数",
            "卖出费用",
            "执行检查",
            "规则检查",
        ]
        self.assertEqual(
            [tracking.cell(1, column).value for column in range(1, 30)],
            expected_headers,
        )
        self.assertEqual(tracking.freeze_panes, "D2")
        self.assertIn("MAXIFS", tracking["O2"].value)
        self.assertIn("MAXIFS", tracking["Q2"].value)
        self.assertIn("MAXIFS", tracking["S2"].value)
        self.assertIn('"触发止损：应全部卖出"', tracking["W2"].value)
        self.assertIn('"执行卖出"', tracking["AB2"].value)
        self.assertIn("计划有效", tracking["AC2"].value)
        self.assertEqual(len(tracking.tables), 1)

    def test_opening_guard_requires_a_valid_stop_plan(self):
        trade = self.workbook["单次交易"]
        self.assertIn("'持仓跟踪'!$A$2:$A$501", trade["G2"].value)
        self.assertIn("禁止开仓：请先填写止损计划", trade["G2"].value)
        share_validations = [
            item
            for item in trade.data_validations.dataValidation
            if "H2:H201" in str(item.sqref)
        ]
        self.assertEqual(len(share_validations), 1)
        self.assertIn("_TrackRule", share_validations[0].formula1)

    def test_trade_expectation_fields_are_required_before_opening(self):
        trade = self.workbook["单次交易"]
        self.assertEqual(
            [trade.cell(1, column).value for column in range(33, 37)],
            [
                "交易预期类型",
                "预期持有周期",
                "预期选择理由",
                "周期匹配检查",
            ],
        )
        self.assertIn("短期博反弹", trade["AJ2"].value)
        self.assertIn("突破冲新高", trade["AJ2"].value)
        self.assertIn("趋势波段", trade["AJ2"].value)
        self.assertIn("周期与交易预期不匹配", trade["AJ2"].value)
        self.assertIn('AG2=""', trade["G2"].value)
        self.assertIn('AH2=""', trade["G2"].value)
        self.assertIn('AI2=""', trade["G2"].value)
        self.assertIn("禁止开仓：交易预期未填写完整", trade["G2"].value)
        self.assertIn("允许开仓（周期需复核）", trade["G2"].value)
        share_validations = [
            item
            for item in trade.data_validations.dataValidation
            if "H2:H201" in str(item.sqref)
        ]
        self.assertEqual(len(share_validations), 1)
        self.assertIn('AG2<>""', share_validations[0].formula1)
        self.assertIn('AH2<>""', share_validations[0].formula1)
        self.assertIn('AI2<>""', share_validations[0].formula1)
        self.assertEqual(next(iter(trade.tables.values())).ref, "A1:AR201")

    def test_three_tranche_columns_and_formulas_use_cumulative_position(self):
        trade = self.workbook["单次交易"]

        self.assertEqual(trade["E1"].value, "第一批买入价")
        self.assertEqual(trade["H1"].value, "第一批买入股数")
        self.assertEqual(trade["I1"].value, "首次买入日期")
        self.assertEqual(trade["O1"].value, "买入总费用")
        self.assertEqual(
            [trade.cell(1, column).value for column in range(37, 45)],
            [
                "第二批买入价",
                "第二批买入股数",
                "第三批买入价",
                "第三批买入股数",
                "实际加权买入价",
                "累计买入股数",
                "当前持仓风险",
                "分仓规则检查",
            ],
        )
        self.assertIn("E2*H2", trade["AO2"].value)
        self.assertIn("AK2*AL2", trade["AO2"].value)
        self.assertIn("AM2*AN2", trade["AO2"].value)
        self.assertIn("H2+IF(AL2", trade["AP2"].value)
        self.assertIn("MAXIFS", trade["AQ2"].value)
        self.assertIn(
            "'持仓跟踪'!$S$2:$S$501",
            trade["AQ2"].value,
        )
        self.assertIn("违规：超过单笔风险上限", trade["AR2"].value)
        self.assertIn("违规：超过账户风险上限", trade["AR2"].value)
        self.assertIn("AO2", trade["U2"].value)
        self.assertIn("AO2", trade["V2"].value)
        self.assertIn("AK2*AL2", trade["X2"].value)
        self.assertIn("AM2*AN2", trade["X2"].value)
        self.assertIn("AK2*AL2", trade["Y2"].value)
        self.assertEqual(next(iter(trade.tables.values())).ref, "A1:AR201")

    def test_tranche_validations_enforce_entry_order_lock_and_risk(self):
        trade = self.workbook["单次交易"]
        validations = trade.data_validations.dataValidation
        first_shares = next(
            item for item in validations if "H2:H201" in str(item.sqref)
        )
        second_price = next(
            item for item in validations if "AK2:AK201" in str(item.sqref)
        )
        second_shares = next(
            item for item in validations if "AL2:AL201" in str(item.sqref)
        )
        third_price = next(
            item for item in validations if "AM2:AM201" in str(item.sqref)
        )
        third_shares = next(
            item for item in validations if "AN2:AN201" in str(item.sqref)
        )

        self.assertIn("AQ2<=C2*D2", first_shares.formula1)
        self.assertIn(
            "SUM($AQ$2:$AQ$201)<=_AccountRiskLimit",
            first_shares.formula1,
        )
        self.assertIn("ISNUMBER(AK2)", second_price.formula1)
        self.assertIn('AK2<>""', second_shares.formula1)
        self.assertIn("MOD(AL2,100)=0", second_shares.formula1)
        self.assertIn("_LockStatus", second_shares.formula1)
        self.assertIn("AQ2<=C2*D2", second_shares.formula1)
        self.assertIn(
            "SUM($AQ$2:$AQ$201)<=_AccountRiskLimit",
            second_shares.formula1,
        )
        self.assertIn('AND(AK2<>"",AL2<>"")', third_price.formula1)
        self.assertIn('AM2<>""', third_shares.formula1)
        self.assertIn("MOD(AN2,100)=0", third_shares.formula1)
        self.assertTrue(second_shares.showErrorMessage)
        self.assertTrue(third_shares.showErrorMessage)
        self.assertIn("_AccountRiskLimit", self.workbook.defined_names)

    def test_tracking_full_sale_uses_cumulative_shares(self):
        tracking = self.workbook["持仓跟踪"]

        self.assertIn(
            "'单次交易'!$AP$2:$AP$201",
            tracking["Z2"].value,
        )
        self.assertNotIn(
            "'单次交易'!$H$2:$H$201",
            tracking["Z2"].value,
        )

    def test_trade_sheet_contains_required_headers_and_guarded_formulas(self):
        ws = self.workbook["单次交易"]
        self.assertEqual(
            [ws.cell(1, column).value for column in range(1, 32)],
            TRADE_HEADERS,
        )
        self.assertEqual(ws["Y1"].value, "单笔仓位收益率")
        self.assertEqual(ws["AE1"].value, "实际账户收益率")
        self.assertIn("ROUNDDOWN", ws["F2"].value)
        self.assertIn("IF(OR(", ws["F2"].value)
        self.assertIn("'账户数据'!$B$9", ws["G2"].value)
        self.assertIn("'账户数据'!$B$7", ws["G2"].value)
        self.assertIn("禁止开仓", ws["G2"].value)
        self.assertIn("'持仓跟踪'!$Z$2:$Z$501", ws["L2"].value)
        self.assertIn('IF(O2="",0,O2)', ws["X2"].value)
        self.assertIn('IF(P2="",0,P2)', ws["X2"].value)
        self.assertIn("E2*H2", ws["X2"].value)
        self.assertNotIn("E2*F2", ws["X2"].value)
        self.assertNotIn("{row}", ws["X2"].value)
        self.assertIn('IF(OR(X2="",AP2=""),"",IF(', ws["Y2"].value)
        self.assertIn('OR(A2=""', ws["AB2"].value)
        self.assertIn("'多次统计数据'!$B$8", ws["AB2"].value)
        self.assertIn("'多次统计数据'!$B$9", ws["AC2"].value)
        self.assertEqual(
            ws["AE2"].value,
            '=IF(OR(X2="",C2="",C2<=0),"",X2/C2)',
        )
        self.assertIn("AE2", ws["Z2"].value)

    def test_input_and_formula_cells_use_distinct_fills(self):
        ws = self.workbook["单次交易"]
        self.assertEqual(ws["C2"].fill.fgColor.rgb, "00DDEBF7")
        self.assertEqual(ws["D2"].fill.fgColor.rgb, "00DDEBF7")
        self.assertEqual(ws["F2"].fill.fgColor.rgb, "00E7E6E6")
        self.assertEqual(ws["G2"].fill.fgColor.rgb, "00E7E6E6")
        self.assertEqual(ws["H2"].fill.fgColor.rgb, "00DDEBF7")
        self.assertEqual(ws["X2"].fill.fgColor.rgb, "00E7E6E6")
        self.assertEqual(ws["AE2"].fill.fgColor.rgb, "00E7E6E6")

    def test_trade_and_reason_logs_are_filterable_and_frozen(self):
        trade = self.workbook["单次交易"]
        reasons = self.workbook["买入理由"]
        self.assertEqual(trade.freeze_panes, "A2")
        self.assertEqual(reasons.freeze_panes, "A2")
        self.assertEqual(len(trade.tables), 1)
        self.assertEqual(len(reasons.tables), 1)
        self.assertGreater(len(trade.data_validations.dataValidation), 5)
        self.assertGreater(len(reasons.data_validations.dataValidation), 2)

    def test_indicator_dropdowns_reference_the_technical_indicator_sheet(self):
        indicators = self.workbook["技术指标"]
        reasons = self.workbook["买入理由"]
        self.assertEqual(indicators["A1"].value, "技术指标")
        self.assertEqual(indicators["A2"].value, "蜡烛图")
        indicator_values = [
            indicators.cell(row, 1).value
            for row in range(2, indicators.max_row + 1)
        ]
        self.assertIn("趋势线", indicator_values)
        self.assertIn("MACD", indicator_values)
        self.assertIn("技术指标列表", self.workbook.defined_names)
        indicator_validations = [
            item
            for item in reasons.data_validations.dataValidation
            if item.formula1 == "=技术指标列表"
        ]
        self.assertEqual(len(indicator_validations), 1)
        self.assertIn("F2:H501", str(indicator_validations[0].sqref))

    def test_statistics_account_and_target_formulas_follow_metric_contract(self):
        stats = self.workbook["多次统计数据"]
        account = self.workbook["账户数据"]
        target = self.workbook["目标收益"]
        self.assertIn("COUNTIF", stats["B3"].value)
        self.assertIn("-AVERAGEIF", stats["B9"].value)
        self.assertIn("'单次交易'!AE2:AE201", stats["B8"].value)
        self.assertIn("'单次交易'!AE2:AE201", stats["B9"].value)
        self.assertIn("B6*B8-B7*B9", stats["B13"].value)
        self.assertIn(
            "SUMPRODUCT(IFERROR(LN(1+'单次交易'!AE2:AE201),0))",
            stats["B14"].value,
        )
        self.assertNotIn(
            "PRODUCT(1+'单次交易'!AE2:AE201)",
            stats["B14"].value,
        )
        self.assertIn("'单次交易'!AO2:AO201", stats["B12"].value)
        self.assertIn("'单次交易'!AP2:AP201", stats["B12"].value)
        self.assertNotIn("'单次交易'!F2:F201", stats["B12"].value)
        self.assertIn("SUM('单次交易'!X2:X201)", account["B3"].value)
        self.assertIn("SUMIFS", account["B8"].value)
        self.assertIn("EOMONTH(TODAY()", account["B8"].value)
        self.assertEqual(account["A6"].value, "当月允许最大亏损比例")
        self.assertEqual(account["A7"].value, "当月允许最大亏损金额")
        self.assertEqual(account["A9"].value, "当前未平仓理论亏损")
        self.assertEqual(account["A10"].value, "当月剩余可开仓风险额度")
        self.assertIn("SUM", account["B9"].value)
        self.assertIn("'单次交易'!AQ2:AQ201", account["B9"].value)
        self.assertEqual(
            account["B10"].value,
            '=IF(OR(B7="",B9=""),"",B7-B9)',
        )
        self.assertEqual(account["A11"].value, "账户实际累计收益率")
        self.assertEqual(
            account["B11"].value,
            '=IF(OR(B2="",B3="",B2<=0),"",B3/B2-1)',
        )
        self.assertIn('"暂不可计算"', target["B7"].value)
        self.assertIn("ROUNDUP", target["B7"].value)
        self.assertIn("LN(1+B3)/LN(1+B6)", target["B7"].value)
        self.assertNotIn("B5*B6", target["B7"].value)

    def test_consecutive_loss_lock_controls_and_opening_guard_are_present(self):
        account = self.workbook["账户数据"]
        trade = self.workbook["单次交易"]

        self.assertEqual(account["A13"].value, "连续亏损锁仓比例")
        self.assertEqual(account["B13"].value, 0.06)
        self.assertEqual(account["A16"].value, "手动解锁截至记录序号")
        self.assertEqual(account["A17"].value, "手动解锁原因")
        self.assertEqual(account["A23"].value, "连续亏损锁仓状态")
        self.assertIn("LOOKUP", account["B14"].value)
        self.assertIn("MAX(B14", account["B18"].value)
        self.assertIn("SUMPRODUCT", account["B20"].value)
        self.assertIn('"解锁信息不完整"', account["B23"].value)
        self.assertIn('"已锁仓"', account["B23"].value)
        self.assertIn("'账户数据'!$B$23", trade["G2"].value)
        self.assertIn("禁止开仓：连续亏损达到上限", trade["G2"].value)
        self.assertIn("_LockStatus", self.workbook.defined_names)
        share_validations = [
            item
            for item in trade.data_validations.dataValidation
            if "H2:H201" in str(item.sqref)
        ]
        self.assertEqual(len(share_validations), 1)
        self.assertIn("_LockStatus", share_validations[0].formula1)

    def test_workbook_has_validations_risk_formatting_and_auto_calculation(self):
        trade = self.workbook["单次交易"]
        self.assertGreater(len(trade.conditional_formatting), 3)
        self.assertEqual(trade["U2"].number_format, "0.00%")
        self.assertEqual(trade["I2"].number_format, "yyyy-mm-dd")
        self.assertEqual(trade["X2"].number_format, '¥#,##0.00;[Red]-¥#,##0.00')
        self.assertEqual(trade["AE2"].number_format, "0.00%")
        self.assertEqual(self.workbook.calculation.calcMode, "auto")
        self.assertTrue(self.workbook.calculation.fullCalcOnLoad)
        self.assertTrue(self.workbook.calculation.forceFullCalc)


class IntegrationTests(unittest.TestCase):
    AS_OF_DATE = date(2026, 7, 24)

    def test_lock_upgrade_preserves_the_existing_trade_history(self):
        source = Path(__file__).resolve().parents[1] / "交易管理系统.xlsx"
        source_values = load_workbook(source, data_only=False)
        with tempfile.TemporaryDirectory() as temporary:
            destination = Path(temporary) / "交易管理系统_V2_连续亏损锁仓.xlsx"

            result = tw.upgrade_workbook_with_consecutive_loss_lock(
                source,
                destination,
            )

            upgraded = load_workbook(result, data_only=False)
            self.assertEqual(upgraded.sheetnames, source_values.sheetnames)
            self.assertEqual(upgraded.active.title, "账户数据")
            self.assertEqual(
                upgraded["单次交易"]["A2"].value,
                source_values["单次交易"]["A2"].value,
            )
            self.assertEqual(
                upgraded["单次交易"]["AF38"].value,
                source_values["单次交易"]["AF38"].value,
            )
            self.assertEqual(
                upgraded["账户数据"]["A23"].value,
                "连续亏损锁仓状态",
            )
            self.assertIsNone(source_values["账户数据"]["A23"].value)

    def test_dynamic_stop_upgrade_preserves_history_and_syncs_future_sales(self):
        source = (
            Path(__file__).resolve().parents[1]
            / "交易管理系统_V2_连续亏损锁仓.xlsx"
        )
        source_values = load_workbook(source, data_only=False)
        source_trade = source_values["单次交易"]
        last_record_row = max(
            row
            for row in range(2, 202)
            if source_trade.cell(row, 1).value not in (None, "")
        )
        with tempfile.TemporaryDirectory() as temporary:
            destination = Path(temporary) / "交易管理系统_V3_动态止损.xlsx"

            result = tw.upgrade_workbook_with_dynamic_stop(source, destination)

            upgraded = load_workbook(result, data_only=False)
            trade = upgraded["单次交易"]
            future_row = last_record_row + 1
            self.assertEqual(upgraded.active.title, "持仓跟踪")
            self.assertEqual(upgraded.sheetnames[1], "持仓跟踪")
            self.assertEqual(trade["K2"].value, source_trade["K2"].value)
            self.assertEqual(trade["M2"].value, source_trade["M2"].value)
            self.assertEqual(trade["AF38"].value, source_trade["AF38"].value)
            self.assertIn(
                "'持仓跟踪'!$Y$2:$Y$501",
                trade.cell(future_row, 11).value,
            )
            self.assertIn(
                "'持仓跟踪'!$Z$2:$Z$501",
                trade.cell(future_row, 12).value,
            )
            self.assertIn(
                "'持仓跟踪'!$H$2:$H$501",
                trade.cell(future_row, 13).value,
            )
            self.assertIn(
                "'持仓跟踪'!$AA$2:$AA$501",
                trade.cell(future_row, 16).value,
            )
            self.assertEqual(
                upgraded["账户数据"]["A23"].value,
                "连续亏损锁仓状态",
            )

    def test_sample_workbook_adds_one_valid_stop_plan_per_trade(self):
        workbook = tw.build_workbook(
            with_sample_data=True,
            as_of_date=self.AS_OF_DATE,
        )
        tracking = workbook["持仓跟踪"]
        expected = tw.generate_sample_transactions(self.AS_OF_DATE)

        self.assertEqual(
            [tracking.cell(row, 1).value for row in range(2, 102)],
            ["止损计划"] * 100,
        )
        self.assertEqual(
            [tracking.cell(row, 2).value for row in range(2, 102)],
            [item["trade_id"] for item in expected],
        )
        self.assertTrue(
            all(
                tracking.cell(row, 5).value
                > workbook["单次交易"].cell(row, 5).value
                for row in range(2, 102)
            )
        )

    def test_trade_expectation_upgrade_preserves_v3_history_and_tracking(self):
        source = (
            Path(__file__).resolve().parents[1]
            / "交易管理系统_V3_动态止损.xlsx"
        )
        source_values = load_workbook(source, data_only=False)
        with tempfile.TemporaryDirectory() as temporary:
            destination = Path(temporary) / "交易管理系统_V4_交易预期.xlsx"

            result = tw.upgrade_workbook_with_trade_expectations(
                source,
                destination,
            )

            upgraded = load_workbook(result, data_only=False)
            trade = upgraded["单次交易"]
            self.assertEqual(upgraded.active.title, "单次交易")
            self.assertEqual(
                trade["A2"].value,
                source_values["单次交易"]["A2"].value,
            )
            self.assertEqual(
                trade["AF38"].value,
                source_values["单次交易"]["AF38"].value,
            )
            self.assertEqual(
                upgraded["持仓跟踪"]["A1"].value,
                source_values["持仓跟踪"]["A1"].value,
            )
            self.assertEqual(trade["AG1"].value, "交易预期类型")
            self.assertEqual(trade["AJ1"].value, "周期匹配检查")
            self.assertIsNone(trade["AG2"].value)
            self.assertEqual(next(iter(trade.tables.values())).ref, "A1:AJ201")

    def test_three_tranche_upgrade_preserves_v4_history(self):
        source = (
            Path(__file__).resolve().parents[1]
            / "交易管理系统_V4_交易预期.xlsx"
        )
        before = load_workbook(source, data_only=False)
        before_trade = before["单次交易"]
        historical_values = {
            coordinate: before_trade[coordinate].value
            for coordinate in (
                "A2",
                "E2",
                "H2",
                "I2",
                "O2",
                "AG2",
                "AJ2",
                "AF38",
            )
        }

        with tempfile.TemporaryDirectory() as temporary:
            destination = Path(temporary) / "交易管理系统_V5_三批买入.xlsx"
            result = tw.upgrade_workbook_with_three_tranche_buying(
                source,
                destination,
            )

            after = load_workbook(result, data_only=False)
            trade = after["单次交易"]
            for coordinate, expected in historical_values.items():
                self.assertEqual(trade[coordinate].value, expected)
            self.assertIsNone(trade["AK2"].value)
            self.assertIsNone(trade["AL2"].value)
            self.assertIsNone(trade["AM2"].value)
            self.assertIsNone(trade["AN2"].value)
            self.assertEqual(trade["E1"].value, "第一批买入价")
            self.assertEqual(trade["AR1"].value, "分仓规则检查")
            self.assertEqual(
                next(iter(trade.tables.values())).ref,
                "A1:AR201",
            )
            self.assertEqual(after.active.title, "单次交易")
            self.assertEqual(after.sheetnames, before.sheetnames)

    def _recalculate_with_libreoffice(self, source: Path, output_dir: Path) -> Path:
        soffice = shutil.which("soffice")
        self.assertIsNotNone(soffice, "LibreOffice is required for formula QA")
        output_dir.mkdir(parents=True)
        profile_dir = output_dir / "lo-profile"
        profile_dir.mkdir()
        command = [
            soffice,
            "--headless",
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(output_dir),
            str(source),
        ]
        completed = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=60,
        )
        self.assertEqual(
            completed.returncode,
            0,
            msg=f"LibreOffice failed: {completed.stdout}\n{completed.stderr}",
        )
        recalculated = output_dir / source.name
        self.assertTrue(
            recalculated.exists(),
            msg=f"LibreOffice did not create {recalculated}: {completed.stdout}",
        )
        return recalculated

    def test_100_trade_sample_workbook_recalculates_to_expected_metrics(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source" / "交易管理系统_测试版.xlsx"
            source.parent.mkdir()
            workbook = tw.build_workbook(
                with_sample_data=True,
                as_of_date=self.AS_OF_DATE,
            )
            workbook.save(source)
            recalculated = self._recalculate_with_libreoffice(
                source,
                root / "recalculated",
            )
            values = load_workbook(recalculated, data_only=True)

            trade = values["单次交易"]
            stats = values["多次统计数据"]
            account = values["账户数据"]
            target = values["目标收益"]

            expected_items = tw.generate_sample_transactions(
                self.AS_OF_DATE,
                count=100,
            )
            expected_trades = tw.sample_trade_metrics(self.AS_OF_DATE)
            for row, item, metric in zip(
                range(2, 102),
                expected_items,
                expected_trades,
            ):
                self.assertEqual(
                    trade.cell(row, 6).value,
                    item["suggested_shares"],
                )
                self.assertEqual(
                    trade.cell(row, 8).value,
                    item["actual_buy_shares"],
                )
                if metric["pnl"] is None:
                    self.assertIsNone(trade.cell(row, 24).value)
                else:
                    self.assertAlmostEqual(
                        trade.cell(row, 24).value,
                        metric["pnl"],
                    )
                    self.assertAlmostEqual(
                        trade.cell(row, 31).value,
                        metric["account_return"],
                    )
            for column in (
                6,
                7,
                12,
                21,
                22,
                23,
                24,
                25,
                26,
                27,
                28,
                29,
                31,
            ):
                self.assertIsNone(
                    trade.cell(102, column).value,
                    msg=f"blank row formula column {column} must stay blank",
                )

            summary = summarize_trades(expected_trades)
            self.assertEqual(stats["B2"].value, summary["completed_count"])
            self.assertEqual(stats["B3"].value, summary["win_count"])
            self.assertEqual(stats["B4"].value, summary["loss_count"])
            self.assertEqual(stats["B5"].value, summary["flat_count"])
            self.assertAlmostEqual(stats["B6"].value, summary["win_rate"])
            self.assertAlmostEqual(stats["B7"].value, summary["loss_rate"])
            self.assertAlmostEqual(stats["B8"].value, summary["average_win"])
            self.assertAlmostEqual(stats["B9"].value, summary["average_loss"])
            self.assertEqual(stats["B10"].value, summary["win_hold_days"])
            self.assertEqual(stats["B11"].value, summary["loss_hold_days"])
            self.assertAlmostEqual(
                stats["B12"].value,
                summary["average_trade_amount"],
            )
            self.assertAlmostEqual(stats["B13"].value, summary["expectancy"])
            self.assertAlmostEqual(
                stats["B14"].value,
                summary["compound_return"],
            )

            current_balance = 100_000 + sum(
                trade["pnl"]
                for trade in expected_trades
                if trade["pnl"] is not None
            )
            self.assertAlmostEqual(account["B3"].value, current_balance)
            self.assertAlmostEqual(
                account["B11"].value,
                current_balance / 100_000 - 1,
            )
            self.assertAlmostEqual(
                stats["B14"].value,
                account["B11"].value,
            )
            expected_monthly_loss = calculate_monthly_loss(
                expected_trades,
                self.AS_OF_DATE,
            )
            self.assertAlmostEqual(
                account["B8"].value,
                expected_monthly_loss,
            )
            self.assertLess(account["B8"].value, 0)
            expected_open_risk = tw.calculate_open_theoretical_loss(
                expected_trades,
            )
            self.assertAlmostEqual(account["B9"].value, expected_open_risk)
            monthly_limit = current_balance * 0.05
            self.assertAlmostEqual(account["B7"].value, monthly_limit)
            self.assertAlmostEqual(
                account["B10"].value,
                monthly_limit - expected_open_risk,
            )
            candidate = expected_items[-1]
            candidate_risk = (
                candidate["suggested_shares"]
                * (candidate["buy_price"] - candidate["stop_price"])
            )
            expected_status = tw.calculate_opening_risk_status(
                expected_open_risk,
                candidate_risk,
                monthly_limit,
            )
            self.assertTrue(trade["G101"].value.startswith(expected_status))
            self.assertEqual(expected_status, "禁止开仓")
            self.assertAlmostEqual(target["B2"].value, current_balance)
            self.assertAlmostEqual(target["B4"].value, current_balance * 0.10)
            expected_count = math.ceil(
                math.log1p(0.10) / math.log1p(summary["expectancy"])
            )
            self.assertEqual(target["B7"].value, expected_count)

    def test_delivery_writer_creates_clean_and_test_workbooks(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            clean, sample = tw.write_workbooks(
                root,
                as_of_date=self.AS_OF_DATE,
            )
            self.assertEqual(clean.name, "交易管理系统.xlsx")
            self.assertEqual(sample.name, "交易管理系统_测试版.xlsx")
            self.assertTrue(clean.exists())
            self.assertTrue(sample.exists())
            clean_wb = load_workbook(clean, data_only=False)
            sample_wb = load_workbook(sample, data_only=False)
            self.assertIsNone(clean_wb["单次交易"]["A2"].value)
            self.assertEqual(sample_wb["单次交易"]["A2"].value, "T2026001")
            self.assertEqual(sample_wb["买入理由"]["D2"].value, "买入")

    def test_progressive_validator_recalculates_after_each_append(self):
        validator = importlib.import_module("progressive_workbook_validation")
        with tempfile.TemporaryDirectory() as temporary:
            result = validator.run_progressive_validation(
                output_dir=temporary,
                as_of_date=self.AS_OF_DATE,
                steps=3,
                quiet=True,
            )
            self.assertEqual(len(result["audit_rows"]), 3)
            self.assertEqual(
                [row["step"] for row in result["audit_rows"]],
                [1, 2, 3],
            )
            self.assertTrue(
                all(row["result"] == "PASS" for row in result["audit_rows"])
            )
            self.assertTrue(result["workbook_path"].exists())
            self.assertTrue(result["report_path"].exists())
            report = result["report_path"].read_text(encoding="utf-8")
            self.assertIn("数据粒度：每行一笔交易", report)
            self.assertIn("交易编号唯一率：100.00%", report)
            self.assertIn("场景覆盖", report)
            final = load_workbook(result["workbook_path"], data_only=True)
            self.assertEqual(final["单次交易"]["A4"].value, "T2026003")
            self.assertIsNone(final["单次交易"]["A5"].value)

    def test_main_writes_deliverables_to_requested_directory(self):
        with tempfile.TemporaryDirectory() as temporary:
            paths = tw.main(
                output_dir=temporary,
                as_of_date=self.AS_OF_DATE,
            )
            self.assertEqual(len(paths), 2)
            self.assertTrue(all(path.exists() for path in paths))


if __name__ == "__main__":
    unittest.main()
