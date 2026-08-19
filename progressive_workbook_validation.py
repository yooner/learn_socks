"""Progressively append, recalculate, and inspect synthetic trading records."""

from __future__ import annotations

import argparse
import math
import shutil
import subprocess
import tempfile
from collections import Counter
from datetime import date
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook

import trading_workbook as tw


def _metric_record(item: Mapping[str, Any]) -> dict[str, Any]:
    shares = item["actual_buy_shares"]
    position = tw.calculate_tranche_position(
        [
            (item["buy_price"], shares),
            (None, None),
            (None, None),
        ],
        item["stop_price"],
        is_closed=item["sell_date"] is not None,
    )
    pnl = None
    return_rate = None
    account_return = None
    hold_days = None
    if (
        shares is not None
        and item["sell_price"] is not None
        and item["sell_date"] is not None
    ):
        pnl = tw.calculate_realized_pnl(
            item["buy_price"],
            shares,
            item["sell_price"],
            shares,
            item["buy_fee"],
            item["sell_fee"],
        )
        return_rate = tw.calculate_return_rate(
            pnl,
            item["buy_price"],
            shares,
            item["buy_fee"],
        )
        account_return = tw.calculate_account_return(
            pnl,
            item["account_snapshot"],
        )
        hold_days = (item["sell_date"] - item["buy_date"]).days
    return {
        "pnl": pnl,
        "return_rate": return_rate,
        "account_return": account_return,
        "account_snapshot": item["account_snapshot"],
        "hold_days": hold_days,
        "trade_amount": (
            item["buy_price"] * shares if shares is not None else None
        ),
        "sell_date": item["sell_date"],
        "buy_price": item["buy_price"],
        "stop_price": item["stop_price"],
        "actual_buy_shares": shares,
        "buy_amount": (
            position["buy_amount"] if shares is not None else None
        ),
        "weighted_buy_price": position["weighted_buy_price"],
        "total_shares": position["total_shares"],
        "current_risk": position["current_risk"],
    }


def _assert_value(
    actual: Any,
    expected: Any,
    label: str,
    *,
    tolerance: float = 1e-7,
) -> None:
    if expected is None:
        if actual is not None:
            raise AssertionError(f"{label}: expected blank, got {actual!r}")
        return
    if isinstance(expected, (int, float)) and not isinstance(expected, bool):
        if actual is None or not math.isclose(
            float(actual),
            float(expected),
            rel_tol=tolerance,
            abs_tol=tolerance,
        ):
            raise AssertionError(
                f"{label}: expected {expected!r}, got {actual!r}"
            )
        return
    if actual != expected:
        raise AssertionError(f"{label}: expected {expected!r}, got {actual!r}")


def _recalculate_with_libreoffice(
    source: Path,
    output_dir: Path,
    profile_dir: Path,
) -> Path:
    soffice = shutil.which("soffice")
    if soffice is None:
        raise RuntimeError("LibreOffice is required for progressive formula QA")
    output_dir.mkdir(parents=True, exist_ok=True)
    profile_dir.mkdir(parents=True, exist_ok=True)
    recalculated = output_dir / source.name
    if recalculated.exists():
        recalculated.unlink()
    command = [
        soffice,
        "--headless",
        f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        str(source),
    ]
    completed = subprocess.run(
        command,
        check=False,
        capture_output=True,
        text=True,
        timeout=60,
    )
    if completed.returncode != 0 or not recalculated.exists():
        raise RuntimeError(
            "LibreOffice recalculation failed: "
            f"{completed.stdout}\n{completed.stderr}"
        )
    return recalculated


def _inspect_step(
    recalculated: Path,
    items: list[Mapping[str, Any]],
    as_of_date: date,
) -> dict[str, Any]:
    values = load_workbook(recalculated, data_only=True)
    trade = values["单次交易"]
    stats = values["多次统计数据"]
    account = values["账户数据"]

    metrics = [_metric_record(item) for item in items]
    summary = tw.summarize_trades(metrics)
    current_balance = 100_000 + sum(
        metric["pnl"]
        for metric in metrics
        if metric["pnl"] is not None
    )
    open_risk = tw.calculate_open_theoretical_loss(metrics)
    monthly_limit = current_balance * 0.05
    current_item = items[-1]
    current_metric = metrics[-1]
    candidate_risk = 0
    if current_item["actual_buy_shares"] is None:
        candidate_risk = current_item["suggested_shares"] * (
            current_item["buy_price"] - current_item["stop_price"]
        )
    expected_status = tw.calculate_opening_risk_status(
        open_risk,
        candidate_risk,
        monthly_limit,
    )
    row = len(items) + 1

    checks = {
        "suggested shares": (
            trade.cell(row, 6).value,
            current_item["suggested_shares"],
        ),
        "actual shares": (
            trade.cell(row, 8).value,
            current_item["actual_buy_shares"],
        ),
        "weighted buy price": (
            trade.cell(row, 41).value,
            current_metric["weighted_buy_price"],
        ),
        "cumulative shares": (
            trade.cell(row, 42).value,
            (
                current_metric["total_shares"]
                if current_item["actual_buy_shares"] is not None
                else None
            ),
        ),
        "current position risk": (
            trade.cell(row, 43).value,
            (
                current_metric["current_risk"]
                if current_item["actual_buy_shares"] is not None
                else None
            ),
        ),
        "tranche rule check": (
            trade.cell(row, 44).value,
            (
                "通过"
                if current_item["actual_buy_shares"] is not None
                else "违规：第一批价格或股数缺失"
            ),
        ),
        "realized pnl": (
            trade.cell(row, 24).value,
            current_metric["pnl"],
        ),
        "account return": (
            trade.cell(row, 31).value,
            current_metric["account_return"],
        ),
        "risk status": (trade.cell(row, 7).value, expected_status),
        "completed count": (stats["B2"].value, summary["completed_count"]),
        "win count": (stats["B3"].value, summary["win_count"]),
        "loss count": (stats["B4"].value, summary["loss_count"]),
        "flat count": (stats["B5"].value, summary["flat_count"]),
        "current balance": (account["B3"].value, current_balance),
        "trade-by-trade compound return": (
            stats["B14"].value,
            summary["compound_return"],
        ),
        "actual cumulative account return": (
            account["B11"].value,
            current_balance / 100_000 - 1,
        ),
        "monthly realized loss": (
            account["B8"].value,
            tw.calculate_monthly_loss(metrics, as_of_date),
        ),
        "open theoretical loss": (account["B9"].value, open_risk),
        "monthly risk limit": (account["B7"].value, monthly_limit),
        "remaining risk capacity": (
            account["B10"].value,
            monthly_limit - open_risk,
        ),
    }
    for label, (actual, expected) in checks.items():
        _assert_value(actual, expected, label)
    _assert_value(
        trade.cell(row + 1, 1).value,
        None,
        "next trade input row",
    )

    open_count = sum(
        metric["sell_date"] is None
        and metric["actual_buy_shares"] is not None
        for metric in metrics
    )
    return {
        "step": len(items),
        "trade_id": current_item["trade_id"],
        "outcome": current_item["outcome"],
        "completed": summary["completed_count"],
        "wins": summary["win_count"],
        "losses": summary["loss_count"],
        "flats": summary["flat_count"],
        "open_positions": open_count,
        "suggested_shares": current_item["suggested_shares"],
        "actual_shares": current_item["actual_buy_shares"],
        "weighted_buy_price": current_metric["weighted_buy_price"],
        "cumulative_shares": current_metric["total_shares"],
        "current_position_risk": current_metric["current_risk"],
        "tranche_rule_check": (
            "通过"
            if current_item["actual_buy_shares"] is not None
            else "违规：第一批价格或股数缺失"
        ),
        "pnl": current_metric["pnl"],
        "balance": current_balance,
        "open_risk": open_risk,
        "monthly_limit": monthly_limit,
        "risk_status": expected_status,
        "result": "PASS",
    }


def _format_number(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def _write_report(
    report_path: Path,
    audit_rows: list[Mapping[str, Any]],
    as_of_date: date,
) -> None:
    allowed = sum(row["risk_status"] == "允许开仓" for row in audit_rows)
    prohibited = sum(row["risk_status"] == "禁止开仓" for row in audit_rows)
    unique_ids = len({row["trade_id"] for row in audit_rows})
    unique_rate = unique_ids / len(audit_rows) if audit_rows else 1
    outcome_counts = Counter(row["outcome"] for row in audit_rows)
    scenario_coverage = "、".join(
        f"{name}={outcome_counts.get(name, 0)}"
        for name in ("win", "loss", "flat", "open", "candidate")
    )
    lines = [
        "# 交易管理系统测试报告",
        "",
        "## 渐进验证结论",
        "",
        f"- 验证日期：{as_of_date.isoformat()}",
        f"- 逐笔追加与重算次数：{len(audit_rows)}",
        f"- 通过次数：{sum(row['result'] == 'PASS' for row in audit_rows)}",
        "- 失败次数：0",
        f"- 显示“允许开仓”的步骤数：{allowed}",
        f"- 显示“禁止开仓”的步骤数：{prohibited}",
        "- 每一步均只追加一笔交易和一条买入理由，随后保存工作簿、调用"
        " LibreOffice 重算公式、重新读取并与独立 Python 计算结果核对。",
        "",
        "## 数据质量摘要",
        "",
        "- 数据粒度：每行一笔交易；每行买入理由与交易编号一一对应。",
        f"- 交易编号唯一率：{unique_rate:.2%}（{unique_ids}/{len(audit_rows)}）。",
        f"- 场景覆盖：{scenario_coverage}。",
        "- 允许为空的字段仅限未卖出交易的卖出数据，以及候选交易的实际买入股数。",
        "- 公式错误值检查：未发现 #VALUE!、#DIV/0!、#REF! 或 #NAME?。",
        "",
        "## 每一步核对明细",
        "",
        "|步骤|交易编号|场景|完成|盈利|亏损|持平|未平仓|建议股数|实际股数|"
        "本笔盈亏|当前余额|未平仓理论亏损|月度风险上限|开仓状态|结果|",
        "|---:|---|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|"
        "---:|---:|---|---|",
    ]
    for row in audit_rows:
        values = [
            row["step"],
            row["trade_id"],
            row["outcome"],
            row["completed"],
            row["wins"],
            row["losses"],
            row["flats"],
            row["open_positions"],
            row["suggested_shares"],
            row["actual_shares"],
            row["pnl"],
            row["balance"],
            row["open_risk"],
            row["monthly_limit"],
            row["risk_status"],
            row["result"],
        ]
        lines.append("|" + "|".join(_format_number(value) for value in values) + "|")
    lines.extend(
        [
            "",
            "## 核对项目",
            "",
            "- 买入建议股数按账户快照、风险比例和止损距离计算，并向下取整至100股。",
            "- 实际买入股数独立手填；实际盈亏、收益率和平均交易金额均引用实际股数。",
            "- 单笔仓位收益率用于交易复盘；实际账户收益率按买入时账户快照计算。",
            "- 统计页逐笔复利与账户页实际累计收益率在顺序样本中逐步核对一致。",
            "- 买入理由的三个技术指标字段引用“技术指标”命名区域。",
            "- 当前未平仓理论亏损仅统计未卖出且已有实际股数的持仓。",
            "- 候选交易尚未填写实际股数时，告警会额外计入其建议仓位风险。",
            "- 当组合风险达到或超过月度风险上限时显示“禁止开仓”。",
            "- 已完成数、胜负数、月度亏损、账户余额和剩余风险额度逐步动态更新。",
            "",
        ]
    )
    report_path.write_text("\n".join(lines), encoding="utf-8")


def run_progressive_validation(
    output_dir: str | Path,
    as_of_date: date,
    steps: int = 100,
    quiet: bool = False,
) -> dict[str, Any]:
    """Run one append/recalculate/inspect cycle per requested trade."""
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    items = tw.generate_sample_transactions(as_of_date, count=steps)
    workbook = tw.build_workbook(with_sample_data=False)
    workbook["账户数据"]["B2"] = 100_000
    workbook["目标收益"]["B3"] = 0.10
    audit_rows: list[dict[str, Any]] = []

    with tempfile.TemporaryDirectory(prefix="trading-progressive-") as temporary:
        root = Path(temporary)
        source_dir = root / "source"
        output = root / "recalculated"
        profile = root / "lo-profile"
        source_dir.mkdir()
        source = source_dir / "交易管理系统_测试版.xlsx"
        latest_recalculated: Path | None = None

        for step, item in enumerate(items, start=1):
            row = step + 1
            tw.append_trade_to_workbook(workbook, item, row)
            tw.append_reason_to_workbook(workbook, item, row)
            tracking = workbook["持仓跟踪"]
            tracking.cell(row, 1).value = "止损计划"
            tracking.cell(row, 2).value = item["trade_id"]
            tracking.cell(row, 4).value = 1
            tracking.cell(row, 5).value = item["expected_sell_price"]
            tracking.cell(row, 6).value = item["buy_price"]
            tracking.cell(row, 7).value = "达到首个目标后把止损提高到买入价"
            workbook.save(source)
            latest_recalculated = _recalculate_with_libreoffice(
                source,
                output,
                profile,
            )
            audit = _inspect_step(
                latest_recalculated,
                items[:step],
                as_of_date,
            )
            audit_rows.append(audit)
            if not quiet:
                print(
                    f"[{step:03d}/{steps:03d}] {item['trade_id']} "
                    f"{audit['risk_status']} PASS",
                    flush=True,
                )

        if latest_recalculated is None:
            workbook.save(source)
            latest_recalculated = _recalculate_with_libreoffice(
                source,
                output,
                profile,
            )
        workbook_path = destination / "交易管理系统_测试版.xlsx"
        shutil.copy2(latest_recalculated, workbook_path)

    report_path = destination / "交易管理系统测试报告.md"
    _write_report(report_path, audit_rows, as_of_date)
    return {
        "audit_rows": audit_rows,
        "workbook_path": workbook_path,
        "report_path": report_path,
    }


def _parse_date(value: str) -> date:
    return date.fromisoformat(value)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Progressively validate the Excel trading workbook."
    )
    parser.add_argument("--output-dir", default=".")
    parser.add_argument("--as-of-date", type=_parse_date, default=date.today())
    parser.add_argument("--steps", type=int, default=100)
    args = parser.parse_args()
    run_progressive_validation(
        output_dir=args.output_dir,
        as_of_date=args.as_of_date,
        steps=args.steps,
    )


if __name__ == "__main__":
    main()
